import streamlit as st
import pandas as pd
import os
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import random

st.set_page_config(page_title="Sistema Educativo", page_icon="📚", layout="wide", initial_sidebar_state="expanded")

GOOGLE_SHEETS_DISPONIBLE = False
try:
    if "gcp_service_account" in st.secrets:
        GOOGLE_SHEETS_DISPONIBLE = True
except Exception:
    GOOGLE_SHEETS_DISPONIBLE = False

def crear_excel_si_no_existe():
    archivo_excel = "sistema_educativo.xlsx"
    if not os.path.exists(archivo_excel):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for trimestre in ["1 Trimestre", "2 Trimestre", "3 Trimestre"]:
            ws = wb.create_sheet(title=trimestre)
            headers = ["Apellido y Nombre", "Curso"] + [f"Mar-{i:02d}" for i in range(1, 32)] + [f"Abr-{i:02d}" for i in range(1, 31)] + [f"May-{i:02d}" for i in range(1, 32)] + ["Nota Asistencia", "Tipo Evaluación", "Eval 1", "Calif 1", "Eval 2", "Calif 2", "Eval 3", "Calif 3", "Eval 4", "Calif 4", "Eval 5", "Calif 5", "Eval 6", "Calif 6", "Nota Final Evaluaciones", "Observaciones"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        wb.save(archivo_excel)
    return archivo_excel

def calcular_nota_asistencia(presentes, totales):
    if totales == 0:
        return 0
    porcentaje = (presentes / totales) * 100
    if porcentaje >= 80:
        return 10   # EX
    elif porcentaje >= 51:
        return 8    # Regular
    else:
        return 5    # M

def calificacion_a_numero(calif):
    try:
        calificaciones = {"M": 2, "R-": 4, "R+": 5, "B": 6, "MB": 8, "EX": 10}
        return calificaciones.get(calif, 0)
    except:
        return 0

def guardar_datos_excel(df, sheet_name, archivo_excel="sistema_educativo.xlsx"):
    """Función para guardar datos en Excel con manejo de errores"""
    try:
        with pd.ExcelWriter(archivo_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return True
    except Exception as e:
        st.error(f"Error guardando en Excel: {e}")
        return False

def generar_backup_detalles():
    """Generar backup detallado en Excel con todas las columnas solicitadas"""
    try:
        archivo_excel = "sistema_educativo.xlsx"
        if not os.path.exists(archivo_excel):
            return False
        
        backup_filename = f"backup_detalles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb_backup = openpyxl.Workbook()
        wb_backup.remove(wb_backup.active)
        
        for trimestre_num in range(1, 4):
            nombre_trimestre = f"{trimestre_num} Trimestre"
            ws_backup = wb_backup.create_sheet(title=nombre_trimestre)
            
            headers_backup = [
                "Fecha y Hora Backup", "Alumno", "Curso", "Trimestre",
                "Asistencia", "Evaluación", "Tipo Evaluación", "Calificación"
            ]
            
            for col_num, header in enumerate(headers_backup, 1):
                cell = ws_backup.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            df_trimestre = pd.read_excel(archivo_excel, sheet_name=nombre_trimestre)
            
            if not df_trimestre.empty:
                row_num = 2
                for idx, row in df_trimestre.iterrows():
                    if pd.notna(row["Apellido y Nombre"]):
                        columnas_asistencia = [col for col in df_trimestre.columns if any(mes in col for mes in ["Mar-", "Abr-", "May-"])]
                        presentes = sum(1 for col in columnas_asistencia if pd.notna(row[col]) and row[col] == "Presente")
                        totales = sum(1 for col in columnas_asistencia if pd.notna(row[col]))
                        porcentaje_asistencia = (presentes / totales * 100) if totales > 0 else 0
                        
                        ws_backup.cell(row=row_num, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                        ws_backup.cell(row=row_num, column=2, value=row["Apellido y Nombre"])
                        ws_backup.cell(row=row_num, column=3, value=row["Curso"])
                        ws_backup.cell(row=row_num, column=4, value=nombre_trimestre)
                        ws_backup.cell(row=row_num, column=5, value=f"{presentes}/{totales} ({porcentaje_asistencia:.1f}%)")
                        ws_backup.cell(row=row_num, column=6, value="")
                        ws_backup.cell(row=row_num, column=7, value="")
                        ws_backup.cell(row=row_num, column=8, value="")
                        row_num += 1
                        
                        for j in range(1, 7):
                            eval_col = f"Eval {j}"
                            calif_col = f"Calif {j}"
                            
                            if pd.notna(row[eval_col]) and pd.notna(row[calif_col]):
                                ws_backup.cell(row=row_num, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                                ws_backup.cell(row=row_num, column=2, value=row["Apellido y Nombre"])
                                ws_backup.cell(row=row_num, column=3, value=row["Curso"])
                                ws_backup.cell(row=row_num, column=4, value=nombre_trimestre)
                                ws_backup.cell(row=row_num, column=5, value="")
                                ws_backup.cell(row=row_num, column=6, value=row[eval_col])
                                ws_backup.cell(row=row_num, column=7, value=row.get("Tipo Evaluación", ""))
                                ws_backup.cell(row=row_num, column=8, value=row[calif_col])
                                row_num += 1
        
        wb_backup.save(backup_filename)
        return True
    except Exception as e:
        st.error(f"Error generando backup: {e}")
        return False

def cargar_datos_desde_sheets(nombre_trimestre):
    """Carga datos desde Google Sheets si no existe el Excel local"""
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        if not GOOGLE_SHEETS_DISPONIBLE:
            return None

        creds_info = {
            "type": st.secrets["gcp_service_account"]["type"],
            "project_id": st.secrets["gcp_service_account"]["project_id"],
            "private_key_id": st.secrets["gcp_service_account"]["private_key_id"],
            "private_key": st.secrets["gcp_service_account"]["private_key"],
            "client_email": st.secrets["gcp_service_account"]["client_email"],
            "client_id": st.secrets["gcp_service_account"]["client_id"],
            "auth_uri": st.secrets["gcp_service_account"]["auth_uri"],
            "token_uri": st.secrets["gcp_service_account"]["token_uri"],
            "auth_provider_x509_cert_url": st.secrets["gcp_service_account"]["auth_provider_x509_cert_url"],
            "client_x509_cert_url": st.secrets["gcp_service_account"]["client_x509_cert_url"],
        }

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = Credentials.from_service_account_info(creds_info, scopes=scopes)
        client = gspread.authorize(creds)

        SPREADSHEET_ID = st.secrets["gcp_service_account"]["sheet_id"]
        spreadsheet = client.open_by_key(SPREADSHEET_ID)

        try:
            ws = spreadsheet.worksheet(nombre_trimestre)
            data = ws.get_all_values()
            if len(data) <= 1:
                return None
            
            headers = data[0]
            rows = data[1:]
            
            df = pd.DataFrame(rows, columns=headers)
            return df
        except Exception:
            return None
    except Exception:
        return None

def sincronizar_google_sheets():
    """Sincroniza todos los datos con Google Sheets de forma legible"""
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        if not GOOGLE_SHEETS_DISPONIBLE:
            return False, "Secrets de Google no configurados"

        creds_info = {
            "type": st.secrets["gcp_service_account"]["type"],
            "project_id": st.secrets["gcp_service_account"]["project_id"],
            "private_key_id": st.secrets["gcp_service_account"]["private_key_id"],
            "private_key": st.secrets["gcp_service_account"]["private_key"],
            "client_email": st.secrets["gcp_service_account"]["client_email"],
            "client_id": st.secrets["gcp_service_account"]["client_id"],
            "auth_uri": st.secrets["gcp_service_account"]["auth_uri"],
            "token_uri": st.secrets["gcp_service_account"]["token_uri"],
            "auth_provider_x509_cert_url": st.secrets["gcp_service_account"]["auth_provider_x509_cert_url"],
            "client_x509_cert_url": st.secrets["gcp_service_account"]["client_x509_cert_url"],
        }

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = Credentials.from_service_account_info(creds_info, scopes=scopes)
        client = gspread.authorize(creds)

        SPREADSHEET_ID = st.secrets["gcp_service_account"]["sheet_id"]
        spreadsheet = client.open_by_key(SPREADSHEET_ID)

        archivo_excel = "sistema_educativo.xlsx"
        if not os.path.exists(archivo_excel):
            return False, "No existe el archivo Excel local. Primero agregá datos simulados."

        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")

        headers = [
            "Última Actualización",
            "Apellido y Nombre",
            "Curso",
            "Trimestre",
            "Días Presentes",
            "Días Ausentes",
            "% Asistencia",
            "Nota Asistencia",
            "Eval 1 — Nombre",
            "Eval 1 — Calif",
            "Eval 2 — Nombre",
            "Eval 2 — Calif",
            "Eval 3 — Nombre",
            "Eval 3 — Calif",
            "Eval 4 — Nombre",
            "Eval 4 — Calif",
            "Eval 5 — Nombre",
            "Eval 5 — Calif",
            "Eval 6 — Nombre",
            "Eval 6 — Calif",
            "Promedio Final"
        ]

        for trimestre_num in range(1, 4):
            nombre_trimestre = f"{trimestre_num} Trimestre"

            try:
                ws = spreadsheet.worksheet(nombre_trimestre)
            except Exception:
                ws = spreadsheet.add_worksheet(title=nombre_trimestre, rows=500, cols=25)

            try:
                df = pd.read_excel(archivo_excel, sheet_name=nombre_trimestre)
            except Exception:
                continue

            if df.empty:
                continue

            rows_data = [headers]

            for _, row in df.iterrows():
                if pd.isna(row.get("Apellido y Nombre")):
                    continue

                columnas_asistencia = [
                    col for col in df.columns
                    if any(mes in str(col) for mes in ["Mar-", "Abr-", "May-"])
                ]
                presentes = sum(
                    1 for col in columnas_asistencia
                    if pd.notna(row.get(col)) and row.get(col) == "Presente"
                )
                totales = sum(
                    1 for col in columnas_asistencia
                    if pd.notna(row.get(col))
                )
                ausentes = totales - presentes
                porcentaje = round((presentes / totales * 100), 1) if totales > 0 else 0
                nota_asistencia = calcular_nota_asistencia(presentes, totales)

                fila = [
                    timestamp,
                    str(row.get("Apellido y Nombre", "")),
                    str(row.get("Curso", "")),
                    nombre_trimestre,
                    presentes,
                    ausentes,
                    f"{porcentaje}%",
                    nota_asistencia,
                ]

                for j in range(1, 7):
                    fila.append(str(row.get(f"Eval {j}", "")))
                    fila.append(str(row.get(f"Calif {j}", "")))

                fila.append(str(row.get("Nota Final Evaluaciones", "")))
                rows_data.append(fila)

            ws.clear()
            ws.update(rows_data, value_input_option="USER_ENTERED")

            ws.format("A1:U1", {
                "backgroundColor": {"red": 0.118, "green": 0.227, "blue": 0.373},
                "textFormat": {
                    "bold": True,
                    "fontSize": 10,
                    "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}
                },
                "horizontalAlignment": "CENTER"
            })

            if len(rows_data) > 1:
                rango_datos = f"A2:U{len(rows_data)}"
                ws.format(rango_datos, {
                    "textFormat": {"fontSize": 10},
                    "horizontalAlignment": "LEFT"
                })

        return True, f"Sincronizado el {timestamp}"

    except Exception as e:
        return False, str(e)

def restaurar_desde_sheets_si_vacio():
    """Si el Excel local está vacío, restaura desde Google Sheets"""
    archivo_excel_local = "sistema_educativo.xlsx"
    try:
        if not GOOGLE_SHEETS_DISPONIBLE:
            return False
        if not os.path.exists(archivo_excel_local):
            return False

        wb_check = openpyxl.load_workbook(archivo_excel_local)
        ws_check = wb_check["1 Trimestre"]
        tiene_datos = ws_check.max_row > 1
        wb_check.close()

        if tiene_datos:
            return False

        # Excel vacío — restaurar desde Sheets
        restaurado = False
        for trimestre_num in range(1, 4):
            nombre_trimestre = f"{trimestre_num} Trimestre"
            df_sheets = cargar_datos_desde_sheets(nombre_trimestre)
            if df_sheets is not None and not df_sheets.empty:
                guardar_datos_excel(df_sheets, nombre_trimestre, archivo_excel_local)
                restaurado = True
        return restaurado
    except Exception:
        return False

def agregar_datos_simulados_completos():
    archivo_excel = "sistema_educativo.xlsx"
    if not os.path.exists(archivo_excel):
        return False
    try:
        # Nombres únicos por curso — sin repetición entre cursos
        alumnos_por_curso = {
            "EF 1A": [
                "García López, Sofía", "Rodríguez Martínez, Ana", "Fernández García, Laura",
                "Sánchez Hernández, María", "López Torres, Carmen",
                "Pérez Díaz, Beatriz", "Gómez Ruiz, Patricia", "Martínez Castro, Fernanda",
                "Romero Vargas, Sofía", "Alvarez Moreno, Isabel"
            ],
            "EF 2A": [
                "Hernández González, Luciana", "Mendoza Silva, Valentina", "Castro Ramos, Isabella",
                "Vargas Morales, Emilia", "Ortiz Ruiz, Camila",
                "Torres Díaz, Florencia", "Ramírez Luna, Agustina", "Vega Castillo, Catalina",
                "Ríos Herrera, Daniela", "Fuentes Mora, Valentina"
            ],
            "EF 1B": [
                "Jiménez Ramos, Andrea", "Morales Vega, Gabriela", "Cruz Fuentes, Alejandra",
                "Reyes Castillo, Natalia", "Herrera Torres, Verónica",
                "Guzmán Pérez, Mariana", "Medina Ríos, Lucía", "Rojas Mendoza, Carolina",
                "Soto Cruz, Paola", "Aguilar Jiménez, Valeria"
            ],
            "EF 2B": [
                "Díaz Morales, Renata", "Muñoz Reyes, Antonella", "Guerrero Herrera, Micaela",
                "Ruiz Guzmán, Julieta", "Flores Medina, Celeste",
                "Núñez Rojas, Bianca", "Molina Soto, Camille", "Silva Aguilar, Abril",
                "Castillo Díaz, Martina", "Moreno Muñoz, Azul"
            ],
            "TD 2A": [
                "Perea Guerrero, Rocío", "Villanueva Ruiz, Pilar", "Ávila Flores, Rebeca",
                "Espinoza Núñez, Ariana", "Contreras Molina, Sofía",
                "Pacheco Silva, Génesis", "Lozano Castillo, Tamara", "Cárdenas Moreno, Almendra",
                "Ibáñez Perea, Nicole", "Vidal Villanueva, Emma"
            ],
            "TD 2B": [
                "Rangel Ávila, Sara", "Pedraza Espinoza, Regina", "Suárez Contreras, Miranda",
                "Cisneros Pacheco, Jimena", "Montes Lozano, Aitana",
                "Delgado Cárdenas, Daniela", "Esquivel Ibáñez, Paula", "Sandoval Vidal, Ivana",
                "Domínguez Rangel, Carla", "Fuentes Pedraza, Lara"
            ]
        }

        evaluaciones_nombres = [
            "Test de Velocidad", "Test de Resistencia", "Test de Flexibilidad",
            "Test de Fuerza", "Test de Coordinación", "Test de Agilidad"
        ]
        calificaciones_posibles = ["B", "MB", "EX", "R+", "R-", "M"]

        for trimestre_num in range(1, 4):
            nombre_trimestre = f"{trimestre_num} Trimestre"
            df_trimestre = pd.read_excel(archivo_excel, sheet_name=nombre_trimestre)

            for curso, alumnos_lista in alumnos_por_curso.items():
                for alumno in alumnos_lista:
                    # Buscar si ya existe este alumno en este curso
                    ya_existe = (
                        not df_trimestre.empty and
                        "Apellido y Nombre" in df_trimestre.columns and
                        "Curso" in df_trimestre.columns and
                        len(df_trimestre[
                            (df_trimestre["Apellido y Nombre"] == alumno) &
                            (df_trimestre["Curso"] == curso)
                        ]) > 0
                    )
                    if ya_existe:
                        continue

                    nueva_fila = {
                        "Apellido y Nombre": alumno,
                        "Curso": curso,
                        "Tipo Evaluación": "Diagnóstico"
                    }

                    for dia in range(1, 32):
                        nueva_fila[f"Mar-{dia:02d}"] = "Presente" if random.random() > 0.2 else "Ausente"
                    for dia in range(1, 31):
                        nueva_fila[f"Abr-{dia:02d}"] = "Presente" if random.random() > 0.2 else "Ausente"
                    for dia in range(1, 32):
                        nueva_fila[f"May-{dia:02d}"] = "Presente" if random.random() > 0.2 else "Ausente"

                    califs_alumno = []
                    for i in range(1, 7):
                        nombre_eval = evaluaciones_nombres[i - 1]
                        calif = random.choice(calificaciones_posibles)
                        nueva_fila[f"Eval {i}"] = nombre_eval
                        nueva_fila[f"Calif {i}"] = calif
                        califs_alumno.append(calificacion_a_numero(calif))

                    nueva_fila["Nota Final Evaluaciones"] = round(
                        sum(califs_alumno) / len(califs_alumno), 1
                    )

                    df_trimestre = pd.concat(
                        [df_trimestre, pd.DataFrame([nueva_fila])],
                        ignore_index=True
                    )

            guardar_datos_excel(df_trimestre, nombre_trimestre, archivo_excel)

        return True
    except Exception as e:
        st.error(f"Error generando datos simulados: {e}")
        return False

def agregar_nuevo_alumno(nombre, curso):
    archivo_excel = "sistema_educativo.xlsx"
    try:
        for trimestre_num in range(1, 4):
            nombre_trimestre = f"{trimestre_num} Trimestre"
            df_trimestre = pd.read_excel(archivo_excel, sheet_name=nombre_trimestre)
            
            nueva_fila = {
                "Apellido y Nombre": nombre,
                "Curso": curso,
                "Tipo Evaluación": "Diagnóstico"
            }
            
            columnas_asistencia = [col for col in df_trimestre.columns if any(mes in col for mes in ["Mar-", "Abr-", "May-"])]
            for col in columnas_asistencia:
                nueva_fila[col] = "Ausente"
            
            for i in range(1, 7):
                nueva_fila[f"Eval {i}"] = ""
                nueva_fila[f"Calif {i}"] = ""
            
            df_trimestre = pd.concat([df_trimestre, pd.DataFrame([nueva_fila])], ignore_index=True)
            guardar_datos_excel(df_trimestre, nombre_trimestre, archivo_excel)
        
        return True
    except Exception as e:
        st.error(f"Error agregando alumno: {e}")
        return False

def obtener_alumnos_disponibles():
    archivo_excel = "sistema_educativo.xlsx"
    try:
        df_1t = pd.read_excel(archivo_excel, sheet_name="1 Trimestre")
        alumnos = df_1t["Apellido y Nombre"].dropna().tolist()
        return ["Todos"] + sorted(alumnos)
    except Exception:
        return ["Todos"]

def crear_grafico_asistencia(presentes, ausentes, nombre_alumno):
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(8, 4))
    labels = ['Presentes', 'Ausentes']
    sizes = [presentes, ausentes]
    colors = ['#2ecc71', '#e74c3c']
    explode = (0.1, 0)
    
    ax.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
    ax.axis('equal')
    plt.title(f'Asistencia - {nombre_alumno}', fontsize=14, fontweight='bold')
    
    return fig

def crear_grafico_evaluaciones(evaluaciones_data, nombre_alumno):
    import matplotlib.pyplot as plt
    if not evaluaciones_data:
        return None
    
    df_grafico = pd.DataFrame(evaluaciones_data)
    if not df_grafico.empty:
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.bar(df_grafico['Evaluación'], df_grafico['Valor Numérico'], color='#3498db')
        ax.set_xlabel('Evaluaciones', fontsize=12)
        ax.set_ylabel('Calificación', fontsize=12)
        ax.set_title(f'Rendimiento en Evaluaciones - {nombre_alumno}', fontsize=14, fontweight='bold')
        ax.set_ylim(0, 10.5)
        plt.xticks(rotation=45, ha='right')
        plt.grid(axis='y', alpha=0.3)
        return fig
    return None

# Inicializar estado
if 'accion_actual' not in st.session_state:
    st.session_state.accion_actual = "dashboard"
if 'nuevas_evaluaciones' not in st.session_state:
    st.session_state.nuevas_evaluaciones = []

archivo_excel = crear_excel_si_no_existe()

# Sidebar con botones de navegación
st.sidebar.markdown("# 📚 Sistema Educativo")
st.sidebar.markdown("---")

if st.sidebar.button("📊 Dashboard", type="primary", key="btn_dashboard"):
    st.session_state.accion_actual = "dashboard"
if st.sidebar.button("📋 Asistencia", type="primary", key="btn_asistencia"):
    st.session_state.accion_actual = "asistencia"
if st.sidebar.button("📝 Evaluaciones", type="primary", key="btn_evaluaciones"):
    st.session_state.accion_actual = "evaluaciones"
if st.sidebar.button("👤 Agregar Alumno", type="primary", key="btn_agregar_alumno"):
    st.session_state.accion_actual = "agregar_alumno"
if st.sidebar.button("📈 Estadística", type="primary", key="btn_estadistica"):
    st.session_state.accion_actual = "estadistica"
if st.sidebar.button("📊 Reporte", type="primary", key="btn_reporte"):
    st.session_state.accion_actual = "reporte"

st.sidebar.markdown("---")

# Guardar y Backup
if st.sidebar.button("💾 Guardar y Backup", type="primary", key="guardar_backup_principal"):
    try:
        if generar_backup_detalles():
            st.sidebar.success("✅ Backup Excel generado!")
        else:
            st.sidebar.error("❌ Error generando backup Excel")
    except Exception as e:
        st.sidebar.error(f"❌ Error Excel: {e}")

    with st.sidebar:
        with st.spinner("Sincronizando Google Sheets..."):
            ok, mensaje = sincronizar_google_sheets()
            if ok:
                st.sidebar.success(f"✅ Google Sheets: {mensaje}")
            else:
                st.sidebar.error(f"❌ Google Sheets: {mensaje}")

# Restaurar desde Google Sheets si Streamlit reinició y perdió los datos
if GOOGLE_SHEETS_DISPONIBLE:
    try:
        if restaurar_desde_sheets_si_vacio():
            st.toast("✅ Datos restaurados desde Google Sheets", icon="☁️")
    except Exception:
        pass

# Mostrar contenido según la acción seleccionada
if st.session_state.accion_actual == "dashboard":
    st.header("📊 Dashboard General")
    
    total_alumnos = 65
    total_evaluaciones = total_alumnos * 6 * 3
    
    col1, col2, col3, col4 = st.columns(4)
    with col1: st.metric("👥 Total Alumnos", total_alumnos, delta="57")
    with col2: st.metric("📊 Promedio Asistencia", "82%", delta="3%")
    with col3: st.metric("📝 Total Evaluaciones", total_evaluaciones, delta=f"+{total_evaluaciones - 720}")
    with col4: st.metric("📈 Promedio General", "7.6", delta="0.2")
    
    st.markdown("---")
    
    resumen_cursos = [
        {"Curso": "EF 1A", "Alumnos": 10, "Asistencia": "85%", "Promedio": "7.8"},
        {"Curso": "EF 2A", "Alumnos": 10, "Asistencia": "78%", "Promedio": "7.2"},
        {"Curso": "EF 1B", "Alumnos": 10, "Asistencia": "82%", "Promedio": "7.5"},
        {"Curso": "EF 2B", "Alumnos": 10, "Asistencia": "80%", "Promedio": "7.6"},
        {"Curso": "TD 2A", "Alumnos": 10, "Asistencia": "76%", "Promedio": "7.2"},
        {"Curso": "TD 2B", "Alumnos": 10, "Asistencia": "80%", "Promedio": "7.6"}
    ]
    df_resumen = pd.DataFrame(resumen_cursos)
    st.dataframe(df_resumen, use_container_width=True)
    
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("📊 Agregar Datos Simulados Completos", type="primary"):
            with st.spinner("Generando datos..."):
                try:
                    if agregar_datos_simulados_completos():
                        st.success("✅ Datos simulados agregados!")
                        st.info("📊 10 alumnos por curso, 6 cursos")
                        st.info("📝 6 evaluaciones por trimestre")
                        st.info("📅 Datos para los 3 trimestres")
                        with st.spinner("Sincronizando con Google Sheets..."):
                            ok, mensaje = sincronizar_google_sheets()
                            if ok:
                                st.success("✅ Google Sheets actualizado!")
                            else:
                                st.warning(f"⚠️ Sheets: {mensaje}")
                        st.rerun()
                    else:
                        st.error("❌ Error generando datos")
                except Exception as e:
                    st.error(f"❌ Error: {e}")
    with col2:
        if st.button("🔄 Actualizar Datos", type="secondary"):
            st.rerun()
    with col3:
        st.write("")

elif st.session_state.accion_actual == "asistencia":
    st.header("📋 Gestión de Asistencia")
    st.markdown("---")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        curso_asistencia = st.selectbox("📂 Seleccionar Curso:", ["Todos", "EF 1A", "EF 2A", "EF 1B", "EF 2B", "TD 2A", "TD 2B"], key="asistencia_curso")
    with col2:
        trimestre_asistencia = st.selectbox("📅 Seleccionar Trimestre:", ["1 Trimestre", "2 Trimestre", "3 Trimestre"], key="asistencia_trimestre")
    with col3:
        fecha_seleccionada = st.date_input("📅 Seleccionar Fecha:", value=datetime.now().date(), key="asistencia_fecha")
    
    st.markdown("---")
    
    st.subheader("📋 Registro de Asistencia - Marcar Rápidamente")
    
    try:
        df_asistencia = pd.read_excel(archivo_excel, sheet_name=trimestre_asistencia)
        if df_asistencia.empty and GOOGLE_SHEETS_DISPONIBLE:
            df_sheets = cargar_datos_desde_sheets(trimestre_asistencia)
            if df_sheets is not None and not df_sheets.empty:
                guardar_datos_excel(df_sheets, trimestre_asistencia, archivo_excel)
                df_asistencia = df_sheets
    except Exception:
        df_asistencia = pd.DataFrame()
        if GOOGLE_SHEETS_DISPONIBLE:
            df_sheets = cargar_datos_desde_sheets(trimestre_asistencia)
            if df_sheets is not None and not df_sheets.empty:
                df_asistencia = df_sheets

    if curso_asistencia != "Todos":
        df_asistencia = df_asistencia[df_asistencia["Curso"] == curso_asistencia]

    if not df_asistencia.empty:
        meses_es = {
            "Jan": "Jan", "Feb": "Feb", "Mar": "Mar", "Apr": "Abr",
            "May": "May", "Jun": "Jun", "Jul": "Jul", "Aug": "Ago",
            "Sep": "Sep", "Oct": "Oct", "Nov": "Nov", "Dec": "Dic"
        }
        mes_en = fecha_seleccionada.strftime("%b")
        mes = meses_es.get(mes_en, mes_en)
        fecha_str = f"{mes}-{fecha_seleccionada.strftime('%d')}"

        if fecha_str not in df_asistencia.columns:
            df_asistencia[fecha_str] = "Ausente"

        # Resumen del día
        st.markdown(f"### 📅 Asistencia del día: **{fecha_seleccionada.strftime('%d/%m/%Y')}**")
        presentes_hoy = sum(
            1 for _, r in df_asistencia.iterrows()
            if pd.notna(r.get("Apellido y Nombre")) and r.get(fecha_str, "Ausente") == "Presente"
        )
        total_hoy = sum(1 for _, r in df_asistencia.iterrows() if pd.notna(r.get("Apellido y Nombre")))
        col_r1, col_r2, col_r3 = st.columns(3)
        with col_r1:
            st.metric("✅ Presentes", presentes_hoy)
        with col_r2:
            st.metric("❌ Ausentes", total_hoy - presentes_hoy)
        with col_r3:
            st.metric("📊 Total", total_hoy)
        st.markdown("---")

        st.write("✅ Marcá la casilla para **Presente** — sin marcar = **Ausente**")

        if 'asistencia_cambios' not in st.session_state:
            st.session_state.asistencia_cambios = {}

        for idx, row in df_asistencia.iterrows():
            if pd.notna(row.get("Apellido y Nombre")):
                key_estado = f"asistencia_{idx}_{fecha_str}"
                
                # Estado actual del alumno para este día
                estado_guardado = row.get(fecha_str, "Ausente")
                if pd.isna(estado_guardado):
                    estado_guardado = "Ausente"
                
                # Usar session_state si ya fue tocado, sino usar el guardado
                if key_estado not in st.session_state:
                    st.session_state[key_estado] = (estado_guardado == "Presente")
                
                es_presente = st.session_state[key_estado]
                
                # Color de fondo según estado
                color_fondo = "rgba(46,204,113,0.12)" if es_presente else "rgba(231,76,60,0.08)"
                color_borde = "#2ecc71" if es_presente else "#e74c3c"
                
                st.markdown(f"""
                <div style='
                    background:{color_fondo};
                    border-left: 4px solid {color_borde};
                    border-radius: 8px;
                    padding: 8px 12px;
                    margin: 4px 0;
                '>
                    <span style='font-weight:600;font-size:15px;color:#2c3e50;'>
                        {row['Apellido y Nombre']}
                    </span>
                    <span style='font-size:12px;color:#7f8c8d;margin-left:8px;'>
                        📂 {row['Curso']}
                    </span>
                </div>
                """, unsafe_allow_html=True)
                
                col_p, col_a = st.columns(2)
                with col_p:
                    if st.button(
                        "✅  PRESENTE",
                        key=f"btn_presente_{idx}_{fecha_str}",
                        type="primary" if es_presente else "secondary",
                        use_container_width=True
                    ):
                        st.session_state[key_estado] = True
                        st.session_state.asistencia_cambios[f"{idx}_{fecha_str}"] = True
                        st.rerun()
                with col_a:
                    if st.button(
                        "❌  AUSENTE",
                        key=f"btn_ausente_{idx}_{fecha_str}",
                        type="primary" if not es_presente else "secondary",
                        use_container_width=True
                    ):
                        st.session_state[key_estado] = False
                        st.session_state.asistencia_cambios[f"{idx}_{fecha_str}"] = False
                        st.rerun()
                
                st.session_state.asistencia_cambios[f"{idx}_{fecha_str}"] = es_presente

        st.markdown("---")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("💾 Guardar Asistencia", type="primary", key="guardar_todos_asistencia"):
                with st.spinner("Guardando asistencia..."):
                    try:
                        cambios_guardados = 0
                        for key, presente in st.session_state.asistencia_cambios.items():
                            if fecha_str in key:
                                idx = int(key.split("_")[0])
                                df_asistencia.at[idx, fecha_str] = "Presente" if presente else "Ausente"
                                cambios_guardados += 1

                        columnas_asist = [c for c in df_asistencia.columns if any(m in c for m in ["Mar-", "Abr-", "May-"])]
                        for idx, row in df_asistencia.iterrows():
                            if pd.notna(row.get("Apellido y Nombre")):
                                p = sum(1 for c in columnas_asist if pd.notna(row.get(c)) and row.get(c) == "Presente")
                                t = sum(1 for c in columnas_asist if pd.notna(row.get(c)))
                                df_asistencia.at[idx, "Nota Asistencia"] = calcular_nota_asistencia(p, t)

                        if guardar_datos_excel(df_asistencia, trimestre_asistencia, archivo_excel):
                            st.success(f"✅ Asistencia del {fecha_seleccionada.strftime('%d/%m/%Y')} guardada — {cambios_guardados} registros")
                            st.session_state.asistencia_cambios = {}
                        else:
                            st.error("❌ Error guardando Excel local")
                            st.stop()

                        with st.spinner("Sincronizando con Google Sheets..."):
                            ok, mensaje = sincronizar_google_sheets()
                            if ok:
                                st.success("✅ Google Sheets actualizado!")
                            else:
                                st.warning(f"⚠️ Excel guardado pero Sheets falló: {mensaje}")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")
        with col2:
            if st.button("🔄 Recargar", type="secondary", key="recargar_asistencia"):
                st.session_state.asistencia_cambios = {}
                st.rerun()
    else:
        st.info("📋 No hay alumnos para mostrar. Primero agregá datos simulados desde el Dashboard.")

elif st.session_state.accion_actual == "evaluaciones":
    st.header("📝 Gestión de Evaluaciones")
    st.markdown("---")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        curso_eval = st.selectbox(
            "📂 Seleccionar Curso:",
            ["Todos", "EF 1A", "EF 2A", "EF 1B", "EF 2B", "TD 2A", "TD 2B"],
            key="eval_curso"
        )
    with col2:
        trimestre_eval = st.selectbox(
            "📅 Seleccionar Trimestre:",
            ["1 Trimestre", "2 Trimestre", "3 Trimestre"],
            key="eval_trimestre"
        )
    with col3:
        fecha_evaluacion = st.date_input(
            "📅 Fecha de Evaluación:",
            value=datetime.now().date(),
            key="eval_fecha"
        )
    
    st.markdown("---")
    
    st.subheader("📝 Sistema de Evaluaciones - Formato Consistente")
    
    with st.expander("➕ Agregar Nueva Evaluación", expanded=False):
        col1, col2, col3, col4 = st.columns([2, 3, 2, 1])

        with col1:
            nuevo_tipo_eval = st.selectbox(
                "Tipo",
                ["Diagnóstico", "Físico", "Técnico", "Desempeño global"],
                key="nuevo_tipo_eval",
                label_visibility="visible"
            )
        with col2:
            nuevo_nombre_eval = st.text_input(
                "Nombre de la evaluación",
                key="nuevo_nombre_eval",
                placeholder="Ej: Test de velocidad",
                label_visibility="visible"
            )
        with col3:
            nueva_calif_eval = st.selectbox(
                "Calificación",
                ["M", "R-", "R+", "B", "MB", "EX"],
                index=3,
                key="nueva_calif_eval",
                label_visibility="visible"
            )
        with col4:
            st.markdown("###")
            if st.button("➕ Agregar", type="primary", key="btn_agregar_nueva_eval"):
                if nuevo_nombre_eval:
                    st.session_state.nuevas_evaluaciones.append({
                        "nombre": nuevo_nombre_eval,
                        "tipo": nuevo_tipo_eval,
                        "calificacion": nueva_calif_eval,
                        "numero": len(st.session_state.nuevas_evaluaciones) + 7
                    })
                    st.success(f"✅ '{nuevo_nombre_eval}' agregada!")
                    st.rerun()
                else:
                    st.error("❌ Ingresá un nombre para la evaluación")
    
    st.markdown("---")
    
    try:
        df_evaluaciones = pd.read_excel(archivo_excel, sheet_name=trimestre_eval)
        if df_evaluaciones.empty and GOOGLE_SHEETS_DISPONIBLE:
            df_sheets = cargar_datos_desde_sheets(trimestre_eval)
            if df_sheets is not None and not df_sheets.empty:
                guardar_datos_excel(df_sheets, trimestre_eval, archivo_excel)
                df_evaluaciones = df_sheets
    except Exception:
        df_evaluaciones = pd.DataFrame()
        if GOOGLE_SHEETS_DISPONIBLE:
            df_sheets = cargar_datos_desde_sheets(trimestre_eval)
            if df_sheets is not None and not df_sheets.empty:
                df_evaluaciones = df_sheets

    if curso_eval != "Todos":
        if not df_evaluaciones.empty and "Curso" in df_evaluaciones.columns:
            df_evaluaciones = df_evaluaciones[df_evaluaciones["Curso"] == curso_eval]

    if not df_evaluaciones.empty:
        # Resumen de evaluaciones cargadas
        total_alumnos_eval = len(df_evaluaciones[df_evaluaciones["Apellido y Nombre"].notna()])
        total_evals = 0
        for _, row in df_evaluaciones.iterrows():
            for j in range(1, 7):
                if pd.notna(row.get(f"Eval {j}")) and pd.notna(row.get(f"Calif {j}")):
                    total_evals += 1

        col_i1, col_i2, col_i3 = st.columns(3)
        with col_i1:
            st.metric("👥 Alumnos cargados", total_alumnos_eval)
        with col_i2:
            st.metric("📝 Evaluaciones totales", total_evals)
        with col_i3:
            st.metric("📅 Trimestre activo", trimestre_eval)
        st.markdown("---")
        
        st.write("📝 **Evaluaciones** - Formato consistente (nombre arriba, nombre abajo)")
        
        if 'evaluaciones_cambios' not in st.session_state:
            st.session_state.evaluaciones_cambios = {}
        
        for idx, row in df_evaluaciones.iterrows():
            if pd.notna(row["Apellido y Nombre"]):
                st.write(f"### **{row['Apellido y Nombre']}** - 📂 {row['Curso']}")
                
                # Encabezado visual
                col_h1, col_h2, col_h3, col_h4, col_h5 = st.columns([1, 3, 2, 2, 1])
                with col_h1:
                    st.caption("Nro.")
                with col_h2:
                    st.caption("Nombre de la evaluación")
                with col_h3:
                    st.caption("Calificación")
                with col_h4:
                    st.caption(f"📅 {fecha_evaluacion.strftime('%d/%m/%Y')}")
                with col_h5:
                    st.caption("Estado")
                
                for j in range(1, 7):
                    eval_col = f"Eval {j}"
                    calif_col = f"Calif {j}"
                    
                    if eval_col in df_evaluaciones.columns and calif_col in df_evaluaciones.columns:
                        col1, col2, col3, col4, col5 = st.columns([1, 3, 2, 2, 1])

                        with col1:
                            tipo_eval = str(row.get('Tipo Evaluación', 'Diagnóstico'))
                            st.markdown(f"**Eval {j}**")
                            st.caption(tipo_eval)

                        with col2:
                            nombre_eval_actual = st.text_input(
                                f"Nombre evaluación {j}",
                                value=str(row.get(eval_col, f"Evaluación {j}")),
                                key=f"eval_nombre_{idx}_{j}",
                                label_visibility="collapsed"
                            )

                        with col3:
                            calificacion_actual = str(row.get(calif_col, "B"))
                            opciones_calif = ["M", "R-", "R+", "B", "MB", "EX"]
                            idx_cal = opciones_calif.index(calificacion_actual) if calificacion_actual in opciones_calif else 3
                            calificacion = st.selectbox(
                                f"Calificación {j}",
                                opciones_calif,
                                index=idx_cal,
                                key=f"eval_calif_{idx}_{j}",
                                label_visibility="collapsed"
                            )

                        with col4:
                            st.caption(f"📅 {fecha_evaluacion.strftime('%d/%m/%Y')}")

                        with col5:
                            iconos_calif = {"EX": "🌟", "MB": "✅", "B": "🔵", "R+": "⚠️", "R-": "🔴", "M": "💔"}
                            st.markdown(f"### {iconos_calif.get(calificacion, '❓')}")

                        st.session_state.evaluaciones_cambios[f"{idx}_{j}"] = {
                            "nombre": nombre_eval_actual,
                            "calificacion": calificacion,
                            "fecha": fecha_evaluacion.strftime("%d/%m/%Y")
                        }
                
                # Nuevas evaluaciones — DENTRO del loop de alumnos
                if st.session_state.nuevas_evaluaciones:
                    for ne in st.session_state.nuevas_evaluaciones:
                        col1, col2, col3, col4, col5 = st.columns([1, 3, 2, 2, 1])

                        with col1:
                            st.markdown(f"**Eval {ne['numero']}**")
                            st.caption(ne['tipo'])

                        with col2:
                            nombre_ne = st.text_input(
                                f"Nombre nueva eval {ne['numero']}",
                                value=ne['nombre'],
                                key=f"eval_nombre_nueva_{ne['numero']}_{idx}",
                                label_visibility="collapsed"
                            )

                        with col3:
                            opciones_ne = ["M", "R-", "R+", "B", "MB", "EX"]
                            idx_ne = opciones_ne.index(ne['calificacion']) if ne['calificacion'] in opciones_ne else 3
                            nueva_cal = st.selectbox(
                                f"Calificación nueva {ne['numero']}",
                                opciones_ne,
                                index=idx_ne,
                                key=f"nueva_eval_cal_{ne['numero']}_{idx}",
                                label_visibility="collapsed"
                            )
                            st.session_state.evaluaciones_cambios[f"nueva_{idx}_{ne['numero']}"] = {
                                "nombre": nombre_ne,
                                "calificacion": nueva_cal,
                                "fecha": fecha_evaluacion.strftime("%d/%m/%Y")
                            }

                        with col4:
                            st.caption(f"📅 {fecha_evaluacion.strftime('%d/%m/%Y')}")

                        with col5:
                            iconos_ne = {"EX": "🌟", "MB": "✅", "B": "🔵", "R+": "⚠️", "R-": "🔴", "M": "💔"}
                            st.markdown(f"### {iconos_ne.get(nueva_cal, '❓')}")
                
                st.markdown("---")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("💾 Guardar Evaluaciones", type="primary", key="guardar_todos_evaluaciones"):
                with st.spinner("Guardando evaluaciones..."):
                    try:
                        cambios_guardados = 0
                        for key, cambios in st.session_state.evaluaciones_cambios.items():
                            try:
                                # Keys normales: "idx_j" (ej: "0_1")
                                # Keys nuevas: "nueva_idx_numero"
                                partes = key.split("_")
                                
                                if key.startswith("nueva_"):
                                    # Evaluación nueva — no se guarda en columnas fijas por ahora
                                    continue
                                
                                if len(partes) >= 2:
                                    idx = int(partes[0])
                                    j = int(partes[1])
                                    eval_col = f"Eval {j}"
                                    calif_col = f"Calif {j}"
                                    if eval_col in df_evaluaciones.columns:
                                        df_evaluaciones.at[idx, eval_col] = cambios["nombre"]
                                        df_evaluaciones.at[idx, calif_col] = cambios["calificacion"]
                                        cambios_guardados += 1
                            except (ValueError, KeyError):
                                continue

                        # Recalcular nota final
                        for idx, row in df_evaluaciones.iterrows():
                            if pd.notna(row.get("Apellido y Nombre")):
                                califs = []
                                for i in range(1, 7):
                                    v = df_evaluaciones.at[idx, f"Calif {i}"]
                                    if pd.notna(v):
                                        califs.append(calificacion_a_numero(v))
                                if califs:
                                    promedio_final = sum(califs) / len(califs)
                                    df_evaluaciones.at[idx, "Nota Final Evaluaciones"] = round(promedio_final, 1)

                        # Guardar Excel local
                        if guardar_datos_excel(df_evaluaciones, trimestre_eval, archivo_excel):
                            st.success(f"✅ {cambios_guardados} evaluaciones guardadas correctamente!")
                            st.session_state.evaluaciones_cambios = {}
                        else:
                            st.error("❌ Error guardando Excel local")
                            st.stop()

                        # Sincronizar con Google Sheets
                        with st.spinner("Sincronizando con Google Sheets..."):
                            ok, mensaje = sincronizar_google_sheets()
                            if ok:
                                st.success("✅ Google Sheets actualizado!")
                            else:
                                st.warning(f"⚠️ Excel guardado pero Sheets falló: {mensaje}")

                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error inesperado: {e}")
        with col2:
            if st.button("🔄 Recargar Datos", type="secondary", key="recargar_evaluaciones"):
                st.session_state.evaluaciones_cambios = {}
                st.rerun()
        with col3:
            calificaciones_contadas = {"M": 0, "R-": 0, "R+": 0, "B": 0, "MB": 0, "EX": 0}
            total_evaluaciones = 0
            
            for _, row in df_evaluaciones.iterrows():
                if pd.notna(row["Apellido y Nombre"]):
                    for i in range(1, 7):
                        calif = df_evaluaciones.at[row.name, f"Calif {i}"]
                        if pd.notna(calif):
                            calificaciones_contadas[calif] += 1
                            total_evaluaciones += 1
            
            if total_evaluaciones > 0:
                st.write("#### 📈 Distribución de Calificaciones")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.write("📈 **Distribución calificaciones:**")
                    for calif, cantidad in calificaciones_contadas.items():
                        st.write(f"- {calif}: {cantidad}")
                with col2:
                    st.write(f"📊 **Total evaluaciones:** {total_evaluaciones}")
                with col3:
                    st.write(f"📈 **Promedio general:** {total_evaluaciones/6:.1f}")
    else:
        st.info("📋 No hay alumnos. Primero agregá datos simulados desde el Dashboard.")

elif st.session_state.accion_actual == "agregar_alumno":
    st.header("👤 Gestión de Alumnos")
    st.markdown("---")

    tab1, tab2, tab3 = st.tabs(["➕ Agregar Alumno", "✏️ Editar / Corregir", "🗑️ Eliminar Alumno"])

    # ── TAB 1: AGREGAR ──────────────────────────────────
    with tab1:
        st.subheader("➕ Agregar Nuevo Alumno")
        col1, col2 = st.columns(2)
        with col1:
            nuevo_nombre = st.text_input("📝 Nombre Completo:", key="nuevo_nombre_alumno", placeholder="Ej: García López, Sofía")
        with col2:
            nuevo_curso = st.selectbox("📂 Curso:", ["EF 1A", "EF 2A", "EF 1B", "EF 2B", "TD 2A", "TD 2B"], key="nuevo_curso_alumno")

        if st.button("➕ Agregar Alumno", type="primary", key="btn_agregar_alumno_def"):
            if nuevo_nombre and nuevo_curso:
                if agregar_nuevo_alumno(nuevo_nombre, nuevo_curso):
                    st.success(f"✅ '{nuevo_nombre}' agregado a {nuevo_curso}")
                    st.rerun()
                else:
                    st.error("❌ Error agregando alumno")
            else:
                st.error("❌ Completá todos los campos")

        st.markdown("---")
        st.write("📋 **Alumnos Actuales:**")
        try:
            df_alumnos = pd.read_excel(archivo_excel, sheet_name="1 Trimestre")
            if not df_alumnos.empty:
                for _, row in df_alumnos.iterrows():
                    if pd.notna(row.get("Apellido y Nombre")):
                        st.write(f"👤 {row['Apellido y Nombre']} — 📂 {row['Curso']}")
        except Exception as e:
            st.error(f"Error: {e}")

    # ── TAB 2: EDITAR / CORREGIR ─────────────────────────
    with tab2:
        st.subheader("✏️ Corregir nombre de alumno")
        st.info("Usá esta sección si te equivocaste al escribir el nombre de un alumno.")

        try:
            df_edit = pd.read_excel(archivo_excel, sheet_name="1 Trimestre")
            alumnos_lista = df_edit["Apellido y Nombre"].dropna().tolist()

            if alumnos_lista:
                alumno_a_editar = st.selectbox(
                    "👤 Seleccioná el alumno a corregir:",
                    alumnos_lista,
                    key="edit_alumno_select"
                )
                nombre_corregido = st.text_input(
                    "📝 Nombre correcto:",
                    value=alumno_a_editar,
                    key="edit_nombre_nuevo"
                )

                if st.button("💾 Guardar Corrección", type="primary", key="btn_guardar_correccion"):
                    if nombre_corregido and nombre_corregido != alumno_a_editar:
                        try:
                            correcciones = 0
                            for trimestre_num in range(1, 4):
                                nombre_trimestre = f"{trimestre_num} Trimestre"
                                df_t = pd.read_excel(archivo_excel, sheet_name=nombre_trimestre)
                                mask = df_t["Apellido y Nombre"] == alumno_a_editar
                                if mask.any():
                                    df_t.loc[mask, "Apellido y Nombre"] = nombre_corregido
                                    guardar_datos_excel(df_t, nombre_trimestre, archivo_excel)
                                    correcciones += 1
                            if correcciones > 0:
                                with st.spinner("Sincronizando..."):
                                    sincronizar_google_sheets()
                                st.success(f"✅ Nombre corregido en {correcciones} trimestres!")
                                st.rerun()
                            else:
                                st.warning("⚠️ No se encontró el alumno")
                        except Exception as e:
                            st.error(f"❌ Error: {e}")
                    else:
                        st.warning("⚠️ El nombre es igual al actual")
            else:
                st.info("No hay alumnos cargados")
        except Exception as e:
            st.error(f"Error: {e}")

        st.markdown("---")
        st.subheader("✏️ Corregir asistencia de un día")
        st.info("Si marcaste mal la asistencia de un día ya guardado, usá esta sección.")

        try:
            df_corr = pd.read_excel(archivo_excel, sheet_name="1 Trimestre")
            alumnos_corr = df_corr["Apellido y Nombre"].dropna().tolist()

            col1, col2, col3 = st.columns(3)
            with col1:
                trimestre_corr = st.selectbox(
                    "📅 Trimestre:",
                    ["1 Trimestre", "2 Trimestre", "3 Trimestre"],
                    key="corr_trimestre"
                )
            with col2:
                alumno_corr = st.selectbox(
                    "👤 Alumno:",
                    alumnos_corr if alumnos_corr else ["Sin datos"],
                    key="corr_alumno"
                )
            with col3:
                fecha_corr = st.date_input(
                    "📅 Fecha a corregir:",
                    value=datetime.now().date(),
                    key="corr_fecha"
                )

            meses_es_c = {
                "Jan": "Jan", "Feb": "Feb", "Mar": "Mar", "Apr": "Abr",
                "May": "May", "Jun": "Jun", "Jul": "Jul", "Aug": "Ago",
                "Sep": "Sep", "Oct": "Oct", "Nov": "Nov", "Dec": "Dic"
            }
            mes_c = meses_es_c.get(fecha_corr.strftime("%b"), fecha_corr.strftime("%b"))
            fecha_str_corr = f"{mes_c}-{fecha_corr.strftime('%d')}"

            nuevo_estado = st.radio(
                "Estado correcto:",
                ["Presente", "Ausente"],
                horizontal=True,
                key="corr_estado"
            )

            if st.button("💾 Corregir Asistencia", type="primary", key="btn_corregir_asistencia"):
                try:
                    df_tc = pd.read_excel(archivo_excel, sheet_name=trimestre_corr)
                    if fecha_str_corr in df_tc.columns:
                        mask = df_tc["Apellido y Nombre"] == alumno_corr
                        if mask.any():
                            df_tc.loc[mask, fecha_str_corr] = nuevo_estado
                            # Recalcular nota asistencia
                            columnas_a = [c for c in df_tc.columns if any(m in c for m in ["Mar-", "Abr-", "May-"])]
                            for i, row in df_tc.iterrows():
                                if pd.notna(row.get("Apellido y Nombre")):
                                    p = sum(1 for c in columnas_a if pd.notna(row.get(c)) and row.get(c) == "Presente")
                                    t = sum(1 for c in columnas_a if pd.notna(row.get(c)))
                                    df_tc.at[i, "Nota Asistencia"] = calcular_nota_asistencia(p, t)
                            guardar_datos_excel(df_tc, trimestre_corr, archivo_excel)
                            with st.spinner("Sincronizando..."):
                                sincronizar_google_sheets()
                            st.success(f"✅ {alumno_corr} — {fecha_str_corr} corregido a {nuevo_estado}!")
                        else:
                            st.warning("⚠️ Alumno no encontrado")
                    else:
                        st.warning(f"⚠️ La fecha {fecha_str_corr} no existe en los registros")
                except Exception as e:
                    st.error(f"❌ Error: {e}")
        except Exception as e:
            st.error(f"Error cargando datos: {e}")

    # ── TAB 3: ELIMINAR ──────────────────────────────────
    with tab3:
        st.subheader("🗑️ Eliminar Alumno")
        st.warning("⚠️ Esta acción elimina al alumno de los 3 trimestres y no se puede deshacer.")

        try:
            df_del = pd.read_excel(archivo_excel, sheet_name="1 Trimestre")
            alumnos_del = df_del["Apellido y Nombre"].dropna().tolist()

            if alumnos_del:
                col1, col2 = st.columns(2)
                with col1:
                    alumno_a_eliminar = st.selectbox(
                        "👤 Seleccioná el alumno a eliminar:",
                        alumnos_del,
                        key="del_alumno_select"
                    )
                with col2:
                    curso_del = st.selectbox(
                        "📂 Curso (para confirmar):",
                        ["EF 1A", "EF 2A", "EF 1B", "EF 2B", "TD 2A", "TD 2B"],
                        key="del_curso_confirm"
                    )

                confirmar = st.text_input(
                    "✍️ Escribí ELIMINAR para confirmar:",
                    key="del_confirmar",
                    placeholder="ELIMINAR"
                )

                if st.button("🗑️ Eliminar Alumno", type="primary", key="btn_eliminar_alumno"):
                    if confirmar == "ELIMINAR":
                        try:
                            eliminados = 0
                            for trimestre_num in range(1, 4):
                                nombre_trimestre = f"{trimestre_num} Trimestre"
                                df_t = pd.read_excel(archivo_excel, sheet_name=nombre_trimestre)
                                mask = (
                                    (df_t["Apellido y Nombre"] == alumno_a_eliminar) &
                                    (df_t["Curso"] == curso_del)
                                )
                                if mask.any():
                                    df_t = df_t[~mask].reset_index(drop=True)
                                    guardar_datos_excel(df_t, nombre_trimestre, archivo_excel)
                                    eliminados += 1
                            if eliminados > 0:
                                with st.spinner("Sincronizando..."):
                                    sincronizar_google_sheets()
                                st.success(f"✅ '{alumno_a_eliminar}' eliminado de {eliminados} trimestres!")
                                st.rerun()
                            else:
                                st.warning("⚠️ No se encontró el alumno en ese curso")
                        except Exception as e:
                            st.error(f"❌ Error: {e}")
                    else:
                        st.error("❌ Escribí ELIMINAR para confirmar")
            else:
                st.info("No hay alumnos cargados")
        except Exception as e:
            st.error(f"Error: {e}")

elif st.session_state.accion_actual == "estadistica":
    st.markdown("""
    <style>
    .stat-card {
        background: linear-gradient(135deg, #1e3a5f 0%, #2d6a9f 100%);
        padding: 16px; border-radius: 10px; margin: 6px 0; color: white;
    }
    .stat-title {
        font-size: 11px; font-weight: 600; letter-spacing: 1px;
        text-transform: uppercase; color: #a8c8e8; margin-bottom: 4px;
    }
    .stat-value {
        font-size: 26px; font-weight: 700; color: white;
    }
    .stat-sub { font-size: 11px; color: #7fb3d3; margin-top: 2px; }
    .seccion-titulo {
        font-size: 14px; font-weight: 700; color: #1e3a5f;
        padding: 8px 0 4px 0;
        border-bottom: 2px solid #2d6a9f; margin-bottom: 10px;
    }
    .badge-ex { background:#1a7a4a; color:white; padding:2px 10px;
        border-radius:20px; font-weight:700; font-size:13px; }
    .badge-mb { background:#2d6a9f; color:white; padding:2px 10px;
        border-radius:20px; font-weight:700; font-size:13px; }
    .badge-b  { background:#4a90d9; color:white; padding:2px 10px;
        border-radius:20px; font-weight:700; font-size:13px; }
    .badge-rp { background:#e8a020; color:white; padding:2px 10px;
        border-radius:20px; font-weight:700; font-size:13px; }
    .badge-rm { background:#d4601a; color:white; padding:2px 10px;
        border-radius:20px; font-weight:700; font-size:13px; }
    .badge-m  { background:#c0392b; color:white; padding:2px 10px;
        border-radius:20px; font-weight:700; font-size:13px; }
    .eval-row {
        display: flex; align-items: center; padding: 8px 4px;
        border-bottom: 1px solid #eee;
    }
    </style>
    """, unsafe_allow_html=True)
    
    st.header("📈 Análisis Estadístico Individual")
    st.markdown("---")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        curso_stats = st.selectbox(
            "📂 Curso:",
            ["Todos", "EF 1A", "EF 2A", "EF 1B", "EF 2B", "TD 2A", "TD 2B"],
            key="stats_curso"
        )
    with col2:
        trimestre_stats = st.selectbox(
            "📅 Trimestre:",
            ["1 Trimestre", "2 Trimestre", "3 Trimestre"],
            key="stats_trimestre"
        )
    with col3:
        alumnos_disponibles = obtener_alumnos_disponibles()
        alumno_stats = st.selectbox(
            "👤 Alumno:",
            alumnos_disponibles,
            key="stats_alumno"
        )

    st.markdown("---")

    try:
        # Intentar leer desde Excel local
        df_stats = pd.read_excel(archivo_excel, sheet_name=trimestre_stats)
        # Si está vacío, intentar desde Sheets
        if df_stats.empty and GOOGLE_SHEETS_DISPONIBLE:
            df_sheets = cargar_datos_desde_sheets(trimestre_stats)
            if df_sheets is not None and not df_sheets.empty:
                guardar_datos_excel(df_sheets, trimestre_stats, archivo_excel)
                df_stats = df_sheets

        if curso_stats != "Todos":
            df_stats = df_stats[df_stats["Curso"] == curso_stats]
        if alumno_stats != "Todos":
            df_stats = df_stats[df_stats["Apellido y Nombre"] == alumno_stats]

        if not df_stats.empty:
            for idx, row in df_stats.iterrows():
                if pd.notna(row.get("Apellido y Nombre")):

                    # Encabezado del alumno
                    st.markdown(f"""
                    <div style='background:linear-gradient(135deg,#1e3a5f,#2d6a9f);
                    padding:16px 20px;border-radius:12px;margin-bottom:16px;'>
                        <div style='font-size:20px;font-weight:700;color:white;'>
                            👤 {row['Apellido y Nombre']}
                        </div>
                        <div style='color:#a8c8e8;font-size:14px;margin-top:4px;'>
                            📂 {row['Curso']} &nbsp;|&nbsp; 📅 {trimestre_stats}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                    # ── ASISTENCIA ──────────────────────────
                    st.markdown('<div class="seccion-titulo">📋 ASISTENCIA</div>', unsafe_allow_html=True)

                    columnas_asistencia = [
                        col for col in df_stats.columns
                        if any(mes in str(col) for mes in ["Mar-", "Abr-", "May-"])
                    ]
                    presentes = sum(
                        1 for col in columnas_asistencia
                        if pd.notna(row.get(col)) and row.get(col) == "Presente"
                    )
                    totales = sum(
                        1 for col in columnas_asistencia
                        if pd.notna(row.get(col))
                    )
                    ausentes = totales - presentes
                    porcentaje = round((presentes / totales * 100), 1) if totales > 0 else 0
                    nota_asistencia = calcular_nota_asistencia(presentes, totales)

                    # Badge de calificación asistencia
                    if nota_asistencia == 10:
                        badge_asist = '<span class="badge-ex">EX — 10</span>'
                        color_asist = "#1a7a4a"
                    elif nota_asistencia == 8:
                        badge_asist = '<span class="badge-b">B — 8</span>'
                        color_asist = "#4a90d9"
                    else:
                        badge_asist = '<span class="badge-m">M — 5</span>'
                        color_asist = "#c0392b"

                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.markdown(f"""<div class="stat-card">
                            <div class="stat-title">Días Presentes</div>
                            <div class="stat-value">{presentes}</div>
                            <div class="stat-sub">de {totales} clases</div>
                        </div>""", unsafe_allow_html=True)
                    with col2:
                        st.markdown(f"""<div class="stat-card">
                            <div class="stat-title">Días Ausentes</div>
                            <div class="stat-value">{ausentes}</div>
                            <div class="stat-sub">faltas registradas</div>
                        </div>""", unsafe_allow_html=True)
                    with col3:
                        st.markdown(f"""<div class="stat-card">
                            <div class="stat-title">% Asistencia</div>
                            <div class="stat-value">{porcentaje}%</div>
                            <div class="stat-sub">{"✅ Regular" if porcentaje >= 80 else "⚠️ Irregular"}</div>
                        </div>""", unsafe_allow_html=True)
                    with col4:
                        st.markdown(f"""<div class="stat-card">
                            <div class="stat-title">Calif. Asistencia</div>
                            <div class="stat-value">{nota_asistencia}</div>
                            <div class="stat-sub">{badge_asist}</div>
                        </div>""", unsafe_allow_html=True)

                    st.markdown("<br>", unsafe_allow_html=True)

                    # ── EVALUACIONES ────────────────────────
                    st.markdown('<div class="seccion-titulo">📝 EVALUACIONES</div>', unsafe_allow_html=True)

                    evaluaciones_detalle = []
                    calificaciones_todas = []

                    # Primero agregar ASISTENCIA como evaluación más
                    evaluaciones_detalle.append({
                        "N°": "A",
                        "Evaluación": "Asistencia al Trimestre",
                        "Calificación": "EX" if nota_asistencia == 10 else ("B" if nota_asistencia == 8 else "M"),
                        "Valor": nota_asistencia,
                        "es_asistencia": True
                    })
                    calificaciones_todas.append(nota_asistencia)

                    # Luego agregar las evaluaciones normales
                    for i in range(1, 7):
                        eval_col = f"Eval {i}"
                        calif_col = f"Calif {i}"
                        if pd.notna(row.get(eval_col)) and pd.notna(row.get(calif_col)):
                            calif_val = str(row.get(calif_col))
                            calif_num = calificacion_a_numero(calif_val)
                            evaluaciones_detalle.append({
                                "N°": i,
                                "Evaluación": str(row.get(eval_col)),
                                "Calificación": calif_val,
                                "Valor": calif_num,
                                "es_asistencia": False
                            })
                            calificaciones_todas.append(calif_num)

                    if evaluaciones_detalle:
                        badges = {
                            "EX": ("badge-ex", "#1a7a4a"),
                            "MB": ("badge-mb", "#2d6a9f"),
                            "B":  ("badge-b",  "#4a90d9"),
                            "R+": ("badge-rp", "#e8a020"),
                            "R-": ("badge-rm", "#d4601a"),
                            "M":  ("badge-m",  "#c0392b")
                        }

                        for ev in evaluaciones_detalle:
                            calif = ev["Calificación"]
                            badge_class, badge_color = badges.get(calif, ("badge-b", "#4a90d9"))
                            es_asist = ev.get("es_asistencia", False)
                            bg_row = "rgba(26,122,74,0.08)" if es_asist else "transparent"

                            col1, col2, col3 = st.columns([1, 6, 2])
                            with col1:
                                st.markdown(
                                    f"<div style='padding:8px 4px;text-align:center;"
                                    f"font-weight:700;color:#1e3a5f;font-size:14px;"
                                    f"background:{bg_row};border-radius:6px;'>#{ev['N°']}</div>",
                                    unsafe_allow_html=True
                                )
                            with col2:
                                icono = "📋 " if es_asist else ""
                                st.markdown(
                                    f"<div style='padding:8px 4px;color:#2c3e50;"
                                    f"font-size:14px;font-weight:{'600' if es_asist else '400'};"
                                    f"background:{bg_row};border-radius:6px;'>"
                                    f"{icono}{ev['Evaluación']}</div>",
                                    unsafe_allow_html=True
                                )
                            with col3:
                                st.markdown(
                                    f"<div style='padding:8px 0px;text-align:left;"
                                    f"background:{bg_row};border-radius:6px;'>"
                                    f"<span class='{badge_class}' style='font-size:13px;'>"
                                    f"{calif} ({ev['Valor']})</span></div>",
                                    unsafe_allow_html=True
                                )

                        st.markdown("<br>", unsafe_allow_html=True)

                        # ── CALIFICACIÓN PROMEDIO FINAL DEL TRIMESTRE ──
                        st.markdown('<div class="seccion-titulo">🏆 CALIFICACIÓN PROMEDIO FINAL DEL TRIMESTRE</div>', unsafe_allow_html=True)

                        promedio_final_trimestre = round(sum(calificaciones_todas) / len(calificaciones_todas), 1)

                        if promedio_final_trimestre >= 9:
                            alerta_color = "#1a7a4a"
                            alerta_borde = "#27ae60"
                            alerta_icono = "🟢"
                            alerta_texto = "Sobresaliente"
                        elif promedio_final_trimestre >= 8:
                            alerta_color = "#1a4a7a"
                            alerta_borde = "#2980b9"
                            alerta_icono = "🔵"
                            alerta_texto = "Muy bueno"
                        elif promedio_final_trimestre >= 7:
                            alerta_color = "#1a5c3a"
                            alerta_borde = "#27ae60"
                            alerta_icono = "🟢"
                            alerta_texto = "Bueno"
                        elif promedio_final_trimestre >= 6:
                            alerta_color = "#7a5a00"
                            alerta_borde = "#f39c12"
                            alerta_icono = "🟡"
                            alerta_texto = "Regular — requiere atención"
                        else:
                            alerta_color = "#7a1a1a"
                            alerta_borde = "#e74c3c"
                            alerta_icono = "🔴"
                            alerta_texto = "Insuficiente — requiere intervención"

                        st.markdown(f"""
                        <div style='
                            border-left: 5px solid {alerta_borde};
                            background: linear-gradient(135deg, {alerta_color}22, {alerta_color}11);
                            padding: 20px 24px;
                            border-radius: 10px;
                            margin: 10px 0;
                            display: flex;
                            align-items: center;
                            gap: 16px;
                        '>
                            <div style='font-size:48px;'>{alerta_icono}</div>
                            <div>
                                <div style='font-size:40px;font-weight:800;color:{alerta_borde};
                                line-height:1;'>{promedio_final_trimestre}</div>
                                <div style='font-size:11px;color:#7f8c8d;margin-top:6px;'>
                                    Promedio de {len(calificaciones_todas)} evaluaciones 
                                    incluyendo asistencia
                                </div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

                    break
        else:
            st.info("📋 No hay datos para analizar. Seleccioná un alumno.")
    except Exception as e:
        st.error(f"Error: {e}")
        st.info("📊 Primero agregá datos simulados desde el Dashboard.")

elif st.session_state.accion_actual == "reporte":
    st.header("📊 Reporte Individual Completo")
    st.markdown("---")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        curso_reporte = st.selectbox("📂 Seleccionar Curso:", ["Todos", "EF 1A", "EF 2A", "EF 1B", "EF 2B", "TD 2A", "TD 2B"], key="reporte_curso")
    with col2:
        trimestre_reporte = st.selectbox("📅 Seleccionar Trimestre:", ["1 Trimestre", "2 Trimestre", "3 Trimestre"], key="reporte_trimestre")
    with col3:
        alumnos_disponibles = obtener_alumnos_disponibles()
        alumno_reporte = st.selectbox("👤 Seleccionar Alumno:", alumnos_disponibles, key="reporte_alumno")
    
    st.markdown("---")
    
    try:
        # Intentar leer desde Excel local
        df_reporte = pd.read_excel(archivo_excel, sheet_name=trimestre_reporte)
        # Si está vacío, intentar desde Sheets
        if df_reporte.empty and GOOGLE_SHEETS_DISPONIBLE:
            df_sheets = cargar_datos_desde_sheets(trimestre_reporte)
            if df_sheets is not None and not df_sheets.empty:
                guardar_datos_excel(df_sheets, trimestre_reporte, archivo_excel)
                df_reporte = df_sheets

        if curso_reporte != "Todos":
            df_reporte = df_reporte[df_reporte["Curso"] == curso_reporte]
        
        if alumno_reporte != "Todos":
            df_reporte = df_reporte[df_reporte["Apellido y Nombre"] == alumno_reporte]
        
        if not df_reporte.empty:
            for idx, row in df_reporte.iterrows():
                if pd.notna(row["Apellido y Nombre"]):
                    st.write(f"# 📊 Reporte Individual: {row['Apellido y Nombre']}")
                    st.write(f"**📂 Curso:** {row['Curso']}")
                    st.write(f"**📅 Trimestre:** {trimestre_reporte}")
                    st.markdown("---")
                    
                    st.write("## 📋 Asistencia Detallada")
                    columnas_asistencia = [col for col in df_reporte.columns if any(mes in col for mes in ["Mar-", "Abr-", "May-"])]
                    asistencia_data = []
                    presentes = 0
                    ausentes = 0
                    
                    for col in columnas_asistencia:
                        if pd.notna(row[col]):
                            estado = row[col]
                            asistencia_data.append({
                                "Fecha": col,
                                "Estado": estado
                            })
                            if estado == "Presente":
                                presentes += 1
                            else:
                                ausentes += 1
                    
                    if asistencia_data:
                        df_asistencia = pd.DataFrame(asistencia_data)
                        st.dataframe(df_asistencia, use_container_width=True)
                        
                        total_dias = presentes + ausentes
                        porcentaje = (presentes / total_dias * 100) if total_dias > 0 else 0
                        nota_asistencia = calcular_nota_asistencia(presentes, total_dias)
                        
                        st.write(f"**📊 Resumen:** {presentes} presentes, {ausentes} ausentes ({porcentaje:.1f}%)")
                        st.write(f"**📊 Nota Asistencia:** {nota_asistencia}")
                    
                    st.markdown("---")
                    st.write("## 📝 Evaluaciones Detalladas")
                    evaluaciones_data = []
                    calificaciones = []
                    
                    for i in range(1, 7):
                        eval_col = f"Eval {i}"
                        calif_col = f"Calif {i}"
                        if pd.notna(row[eval_col]) and pd.notna(row[calif_col]):
                            calif_num = calificacion_a_numero(row[calif_col])
                            evaluaciones_data.append({
                                "Evaluación": row[eval_col],
                                "Calificación": row[calif_col],
                                "Valor Numérico": calif_num
                            })
                            calificaciones.append(calif_num)
                    
                    if evaluaciones_data:
                        df_evaluaciones = pd.DataFrame(evaluaciones_data)
                        st.dataframe(df_evaluaciones, use_container_width=True)
                        
                        if calificaciones:
                            promedio_eval_num = sum(calificaciones) / len(calificaciones)
                            max_calificacion = max(calificaciones)
                            min_calificacion = min(calificaciones)
                            
                            st.write(f"**📊 Promedio Evaluaciones:** {promedio_eval_num:.1f}")
                            st.write(f"**📊 Calificación Más Alta:** {max_calificacion}")
                            st.write(f"**📊 Calificación Más Baja:** {min_calificacion}")
                            
                            st.markdown("---")
                            st.write("#### 📈 Gráfico de Desempeño")
                            fig = crear_grafico_evaluaciones(evaluaciones_data, row["Apellido y Nombre"])
                            if fig is not None:
                                st.pyplot(fig)
                            else:
                                st.info("No hay datos suficientes para el gráfico")
                    
                    st.markdown("---")
                    st.markdown("## 🏆 Calificación Promedio Final del Trimestre")

                    nota_asistencia_num = calcular_nota_asistencia(presentes, total_dias)
                    promedio_eval_num = round(sum(calificaciones) / len(calificaciones), 1) if calificaciones else 0
                    todas_califs = [nota_asistencia_num] + calificaciones
                    promedio_final_trimestre = round(sum(todas_califs) / len(todas_califs), 1)

                    if promedio_final_trimestre >= 9:
                        alerta_borde = "#27ae60"
                        alerta_color = "#1a7a4a"
                        alerta_icono = "🟢"
                    elif promedio_final_trimestre >= 8:
                        alerta_borde = "#2980b9"
                        alerta_color = "#1a4a7a"
                        alerta_icono = "🔵"
                    elif promedio_final_trimestre >= 7:
                        alerta_borde = "#27ae60"
                        alerta_color = "#1a5c3a"
                        alerta_icono = "🟢"
                    elif promedio_final_trimestre >= 6:
                        alerta_borde = "#f39c12"
                        alerta_color = "#7a5a00"
                        alerta_icono = "🟡"
                    else:
                        alerta_borde = "#e74c3c"
                        alerta_color = "#7a1a1a"
                        alerta_icono = "🔴"

                    st.markdown(f"""
                    <div style='
                        border-left: 5px solid {alerta_borde};
                        background: linear-gradient(135deg, {alerta_color}22, {alerta_color}11);
                        padding: 20px 24px;
                        border-radius: 10px;
                        margin: 10px 0;
                        display: flex;
                        align-items: center;
                        gap: 16px;
                    '>
                        <div style='font-size:48px;'>{alerta_icono}</div>
                        <div>
                            <div style='font-size:40px;font-weight:800;
                            color:{alerta_borde};line-height:1;'>
                                {promedio_final_trimestre}
                            </div>
                            <div style='font-size:11px;color:#7f8c8d;margin-top:6px;'>
                                Asistencia ({nota_asistencia_num}) + 
                                Evaluaciones ({promedio_eval_num}) — 
                                {len(todas_califs)} items promediados
                            </div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    break
        else:
            st.info("📋 No hay datos para mostrar")
    except Exception as e:
        st.error(f"Error: {e}")
        st.info("📊 Agrega datos simulados para probar")
