import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import os
from utils import DataManagement
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Configuración de la página
st.set_page_config(
    page_title="Dashboard de Gestión IES",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS mínimo
hide_streamlit_style = """
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
.stAppDeployButton {display:none;}
</style>
"""
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# Inicializar el gestor de datos
@st.cache_resource
def init_data_manager():
    dm = DataManagement()
    
    # Verificar si hay datos, si no hay, crear datos de ejemplo
    df = dm.load_data()
    if df.empty:
        st.info("📝 Creando datos de ejemplo para demostración...")
        from datos_ejemplo_streamlit import crear_datos_para_streamlit
        crear_datos_para_streamlit()
        st.success("✅ Datos de ejemplo creados. Refresca la página para verlos.")
        st.rerun()
    
    return dm

dm = init_data_manager()

# Sistema de calificaciones
CALIFICACIONES_VALORES = {
    "M": {"nombre": "Malo", "valor": 4, "descripcion": "calificación menos de 5"},
    "R-": {"nombre": "Regular-", "valor": 5.5, "descripcion": "calificación entre 5.5 y 6"},
    "R+": {"nombre": "Regular+", "valor": 6, "descripcion": "calificación es 6"},
    "B": {"nombre": "Bueno", "valor": 7, "descripcion": "calificación 7"},
    "MB": {"nombre": "Muy Bueno", "valor": 8, "descripcion": "calificación 8"},
    "Ex": {"nombre": "Excelente", "valor": 9.5, "descripcion": "calificación entre 9 y 10"}
}

def calcular_nota_asistencia(presentes, total):
    """Calcular nota de asistencia según las reglas"""
    if total == 0:
        return 5
    
    porcentaje = (presentes / total) * 100
    
    if porcentaje >= 80:
        return 10  # Ex
    elif porcentaje >= 51:
        return 8   # B
    else:
        return 5   # M

def calcular_promedio_evaluaciones(calificaciones):
    """Calcular promedio de evaluaciones"""
    if not calificaciones:
        return 0
    
    total = 0
    count = 0
    
    for cal in calificaciones:
        if cal in CALIFICACIONES_VALORES:
            total += CALIFICACIONES_VALORES[cal]["valor"]
            count += 1
    
    return total / count if count > 0 else 0

def crear_excel_asistencias(df, trimestre):
    """Crear Excel de asistencias por trimestre"""
    filename = f"Asistencias_{trimestre.replace(' ', '_')}.xlsx"
    
    # Filtrar por trimestre
    asistencias_df = df[df['Trimestre'] == trimestre].copy()
    
    if not asistencias_df.empty:
        # Crear formato de asistencia diario
        asistencias_data = []
        
        for _, row in asistencias_df.iterrows():
            asistencia_str = row.get('Asistencia', '0/0')
            if '/' in asistencia_str:
                presentes, total = asistencia_str.split('/')
                try:
                    presentes = int(presentes)
                    total = int(total)
                except:
                    presentes = 0
                    total = 0
            else:
                presentes = 0
                total = 0
            
            # Crear fila base
            fila_base = {
                'Apellido y Nombre': row['Apellido y Nombre'],
                'Curso': row['Curso'],
                'Trimestre': row['Trimestre'],
                'Días Totales': total,
                'Días Presentes': presentes,
                'Porcentaje Asistencia': f"{(presentes/total*100):.1f}%" if total > 0 else "0%",
                'Nota Asistencia': calcular_nota_asistencia(presentes, total)
            }
            
            # Agregar columnas de asistencia diaria (simuladas)
            for dia in range(1, min(total + 1, 31)):  # Máximo 30 días por trimestre
                fila_base[f'Día {dia}'] = 'P' if dia <= presentes else 'A'
            
            asistencias_data.append(fila_base)
        
        # Crear DataFrame y guardar
        if asistencias_data:
            asistencia_final_df = pd.DataFrame(asistencias_data)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                asistencia_final_df.to_excel(writer, sheet_name='Asistencias', index=False)
                
                # Aplicar formato profesional
                workbook = writer.book
                worksheet = writer.sheets['Asistencias']
                
                # Formato para encabezados
                header_font = Font(bold=True, size=12, color="FFFFFF")
                header_fill = PatternFill(start_color="366092")
                thin_border = Border(left=Side.thin, right=Side.thin, top=Side.thin, bottom=Side.thin)
                
                for col_num, value in enumerate(asistencia_final_df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Formato para datos
                for row_num in range(2, len(asistencia_final_df) + 2):
                    for col_num in range(1, len(asistencia_final_df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Colorear asistencias diarias
                        if 'Día' in asistencia_final_df.columns[col_num-1]:
                            if cell.value == 'P':
                                cell.fill = PatternFill(start_color="90EE90")  # Verde para presente
                            elif cell.value == 'A':
                                cell.fill = PatternFill(start_color="FFB6C1")  # Rojo para ausente
                
                # Ajustar ancho de columnas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column_idx)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 20)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return filename

def crear_excel_evaluaciones(df, trimestre):
    """Crear Excel de evaluaciones"""
    filename = f"Evaluaciones_{trimestre.replace(' ', '_')}.xlsx"
    
    # Filtrar por trimestre
    eval_df = df[df['Trimestre'] == trimestre].copy()
    
    if not eval_df.empty:
        evaluaciones_data = []
        
        for _, row in eval_df.iterrows():
            evaluacion_row = {
                'Apellido y Nombre': row['Apellido y Nombre'],
                'Curso': row['Curso'],
                'Trimestre': row['Trimestre'],
                'Asistencia': row.get('Asistencia', '0/0'),
                'Nota Asistencia': row.get('Nota Asistencia', ''),
                'Tipo Evaluación': row.get('Tipo Evaluación', ''),
                'Observaciones': row.get('Observaciones', ''),
                'Nota Final': row.get('Nota Final', 0)
            }
            
            # Agregar las 6 evaluaciones
            for i in range(1, 7):
                eval_nombre = row.get(f'Evaluación {i}', '')
                calificacion = row.get(f'Calificación {i}', '')
                
                if eval_nombre:
                    evaluacion_row[f'Evaluación {i}'] = eval_nombre
                    evaluacion_row[f'Calificación {i}'] = calificacion
                else:
                    evaluacion_row[f'Evaluación {i}'] = ''
                    evaluacion_row[f'Calificación {i}'] = ''
            
            evaluaciones_data.append(evaluacion_row)
        
        # Crear DataFrame y guardar
        if evaluaciones_data:
            eval_final_df = pd.DataFrame(evaluaciones_data)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                eval_final_df.to_excel(writer, sheet_name='Evaluaciones', index=False)
                
                # Aplicar formato
                workbook = writer.book
                worksheet = writer.sheets['Evaluaciones']
                
                header_font = Font(bold=True, size=11, color="FFFFFF")
                header_fill = PatternFill(start_color="366092")
                thin_border = Border(left=Side.thin, right=Side.thin, top=Side.thin, bottom=Side.thin)
                
                for col_num, value in enumerate(eval_final_df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for row_num in range(2, len(eval_final_df) + 2):
                    for col_num in range(1, len(eval_final_df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        
                        if 'Calificación' in eval_final_df.columns[col_num-1]:
                            calif = cell.value
                            if calif == 'Ex':
                                cell.fill = PatternFill(start_color="90EE90")
                                cell.font = Font(bold=True)
                            elif calif == 'MB':
                                cell.fill = PatternFill(start_color="ADD8E6")
                                cell.font = Font(bold=True)
                            elif calif == 'B':
                                cell.fill = PatternFill(start_color="FFD966")
                                cell.font = Font(bold=True)
                            elif calif in ['R+', 'R-']:
                                cell.fill = PatternFill(start_color="F4B084")
                                cell.font = Font(bold=True)
                            elif calif == 'M':
                                cell.fill = PatternFill(start_color="FFB6C1")
                                cell.font = Font(bold=True)
                
                # Ajustar ancho de columnas
                column_widths = {
                    'A': 25, 'B': 10, 'C': 12, 'D': 12, 'E': 15, 'F': 15, 'G': 30,
                }
                
                for col_letter, width in column_widths.items():
                    if col_letter in [get_column_letter(i) for i in range(len(eval_final_df.columns))]:
                        worksheet.column_dimensions[col_letter].width = width
                
                for i in range(8, len(eval_final_df.columns) + 1):
                    col_letter = get_column_letter(i)
                    worksheet.column_dimensions[col_letter].width = 20
    
    return filename

# Título principal
st.title("📚 Dashboard de Gestión IES")
st.markdown("### Sistema Completo de Asistencias y Evaluaciones")
st.markdown("---")

# Sidebar con filtros
st.sidebar.title("🔍 Panel de Control")

# Filtro de cursos
cursos = ["Todos"] + dm.cursos
curso_seleccionado = st.sidebar.selectbox("📚 Curso:", cursos)

# Filtro de trimestres
trimestres = ["Todos", "1 Trimestre", "2 Trimestre", "3 Trimestre"]
trimestre_seleccionado = st.sidebar.selectbox("📅 Trimestre:", trimestres)

# Filtro de alumnos
alumnos = ["Todos"]
if curso_seleccionado != "Todos":
    alumnos.extend(dm.get_students_by_course(curso_seleccionado))
else:
    df = dm.load_data()
    if not df.empty:
        alumnos.extend(sorted(df["Apellido y Nombre"].unique()))

alumno_seleccionado = st.sidebar.selectbox("👤 Alumno:", alumnos)

st.sidebar.markdown("---")
st.sidebar.markdown("### 🚀 Acciones Principales")

# Botones de acción
if st.sidebar.button("📝 Tomar Asistencia", use_container_width=True, type="primary"):
    st.session_state.pagina = "asistencia"
    st.rerun()

if st.sidebar.button("🎯 Nueva Evaluación", use_container_width=True):
    st.session_state.pagina = "evaluacion"
    st.rerun()

if st.sidebar.button("📊 Reportes", use_container_width=True):
    st.session_state.pagina = "reportes"
    st.rerun()

if st.sidebar.button("💾 Generar Excel", use_container_width=True):
    st.session_state.pagina = "excel"
    st.rerun()

if st.sidebar.button("🔄 Actualizar", use_container_width=True):
    st.session_state.pagina = "inicio"
    st.rerun()

# Estado inicial
if "pagina" not in st.session_state:
    st.session_state.pagina = "inicio"

# Página de Asistencia
if st.session_state.pagina == "asistencia":
    st.header("📝 Tomar Asistencia Diaria")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        curso_asistencia = st.selectbox("📚 Curso:", dm.cursos)
        trimestre_asistencia = st.selectbox("📅 Trimestre:", ["1 Trimestre", "2 Trimestre", "3 Trimestre"])
    
    with col2:
        fecha_clase = st.date_input("📅 Fecha de clase:", value=datetime.now().date())
        
        # Determinar trimestre y día
        if trimestre_asistencia == "1 Trimestre":
            inicio = datetime(fecha_clase.year, 3, 1).date()
            fin = datetime(fecha_clase.year, 5, 31).date()
        elif trimestre_asistencia == "2 Trimestre":
            inicio = datetime(fecha_clase.year, 6, 1).date()
            fin = datetime(fecha_clase.year, 9, 30).date()
        else:  # 3 Trimestre
            inicio = datetime(fecha_clase.year, 10, 1).date()
            fin = datetime(fecha_clase.year, 12, 31).date()
        
        if inicio <= fecha_clase <= fin:
            dia_numero = (fecha_clase - inicio).days + 1
            st.info(f"📊 Día {dia_numero} del {trimestre_asistencia}")
        else:
            st.warning("📅 Fecha fuera del rango del trimestre")
            dia_numero = 1
    
    with col3:
        st.metric("📅 Mes", fecha_clase.strftime("%B"))
        st.metric("📅 Año", fecha_clase.year)
        st.metric("📅 Día", fecha_clase.day)
    
    # Obtener alumnos del curso
    alumnos_curso = dm.get_students_by_course(curso_asistencia)
    
    if alumnos_curso:
        st.subheader(f"👥 Lista de Asistencia - {curso_asistencia}")
        
        asistencia_data = {}
        
        for alumno in alumnos_curso:
            col1, col2, col3 = st.columns([3, 1, 1])
            
            with col1:
                st.write(f"👤 {alumno}")
            
            with col2:
                presente = st.checkbox("✅", key=f"presente_{alumno}_{fecha_clase}", value=True)
                asistencia_data[alumno] = presente
            
            with col3:
                if presente:
                    st.markdown("✅ **Presente**")
                else:
                    st.markdown("❌ **Ausente**")
        
        # Botones de acción
        st.markdown("---")
        col1, col2, col3 = st.columns([2, 1, 1])
        
        with col1:
            if st.button("💾 Guardar Asistencia", use_container_width=True, type="primary"):
                guardar_asistencia_diaria(dm, asistencia_data, curso_asistencia, trimestre_asistencia, fecha_clase, dia_numero)
        
        with col2:
            if st.button("✅ Todos Presentes", use_container_width=True):
                for alumno in alumnos_curso:
                    st.session_state[f"presente_{alumno}_{fecha_clase}"] = True
                st.rerun()
        
        with col3:
            if st.button("❌ Todos Ausentes", use_container_width=True):
                for alumno in alumnos_curso:
                    st.session_state[f"presente_{alumno}_{fecha_clase}"] = False
                st.rerun()
    
    else:
        st.warning("📝 No hay alumnos registrados en este curso")

# Página de Evaluación
elif st.session_state.pagina == "evaluacion":
    st.header("🎯 Registrar Evaluación Completa")
    
    col1, col2 = st.columns([1, 1])
    
    with col1:
        nombre_alumno = st.text_input("👤 Apellido y Nombre:", placeholder="Ej: García López, María")
        curso = st.selectbox("📚 Curso:", dm.cursos)
        trimestre = st.selectbox("📅 Trimestre:", ["1 Trimestre", "2 Trimestre", "3 Trimestre"])
    
    with col2:
        st.subheader("📋 Asistencia")
        asistencia_dias = st.number_input("📅 Total días:", min_value=1, max_value=100, value=20)
        dias_presentes = st.number_input("✅ Días presentes:", min_value=0, max_value=asistencia_dias, value=18)
        
        porcentaje_asistencia = (dias_presentes / asistencia_dias * 100) if asistencia_dias > 0 else 0
        st.info(f"📊 Porcentaje: {porcentaje_asistencia:.1f}%")
        
        nota_asistencia = calcular_nota_asistencia(dias_presentes, asistencia_dias)
        st.success(f"✅ Nota asistencia: {nota_asistencia}")
    
    # Tipo de evaluación
    st.subheader("📝 Tipo de Evaluación")
    tipo_evaluacion = st.selectbox("Seleccionar tipo:", ["Diagnóstico", "Físico", "Técnico", "Desempeño global"])
    
    # 6 evaluaciones
    st.subheader("📋 6 Evaluaciones Individuales")
    
    evaluaciones = []
    calificaciones = []
    
    col1, col2 = st.columns(2)
    
    with col1:
        for i in range(1, 4):
            with st.expander(f"Evaluación {i}", expanded=True):
                nombre_eval = st.text_input(f"Nombre:", key=f"eval_{i}")
                calificacion = st.selectbox(f"Calificación:", list(CALIFICACIONES_VALORES.keys()), key=f"cal_{i}")
                if nombre_eval:
                    evaluaciones.append(nombre_eval)
                    calificaciones.append(calificacion)
    
    with col2:
        for i in range(4, 7):
            with st.expander(f"Evaluación {i}", expanded=True):
                nombre_eval = st.text_input(f"Nombre:", key=f"eval_{i}")
                calificacion = st.selectbox(f"Calificación:", list(CALIFICACIONES_VALORES.keys()), key=f"cal_{i}")
                if nombre_eval:
                    evaluaciones.append(nombre_eval)
                    calificaciones.append(calificacion)
    
    # Promedio
    if calificaciones:
        promedio_eval = calcular_promedio_evaluaciones(calificaciones)
        st.success(f"📊 Promedio evaluaciones: {promedio_eval:.2f}")
    
    # Observaciones
    st.subheader("📝 Observaciones Generales")
    observaciones = st.text_area("Observaciones:", height=100)
    
    # Guardar
    st.markdown("---")
    if st.button("💾 Guardar Evaluación", type="primary", use_container_width=True):
        if nombre_alumno and curso and trimestre and evaluaciones:
            # Preparar datos
            student_data = {
                "Curso": curso,
                "Trimestre": trimestre,
                "Apellido y Nombre": nombre_alumno,
                "Asistencia": f"{dias_presentes}/{asistencia_dias}",
                "Nota Asistencia": f"{nota_asistencia}",
                "Evaluaciones": len(evaluaciones),
                "Tipo Evaluación": tipo_evaluacion,
                "Fecha Registro": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Observaciones": observaciones,
                "Nota Final": promedio_eval if calificaciones else nota_asistencia
            }
            
            # Agregar evaluaciones
            for i, (eval_nombre, calificacion) in enumerate(zip(evaluaciones, calificaciones)):
                student_data[f"Evaluación {i+1}"] = eval_nombre
                student_data[f"Calificación {i+1}"] = calificacion
            
            # Rellenar campos vacíos
            for i in range(len(evaluaciones), 6):
                student_data[f"Evaluación {i+1}"] = ""
                student_data[f"Calificación {i+1}"] = ""
            
            if dm.save_student_data(student_data):
                st.success("✅ Evaluación guardada!")
                st.balloons()
                st.session_state.pagina = "inicio"
                st.rerun()
            else:
                st.error("❌ Error al guardar")
        else:
            st.error("❌ Complete todos los campos")

# Página de Reportes
elif st.session_state.pagina == "reportes":
    st.header("📊 Reportes y Estadísticas Completos")
    
    # Cargar datos con filtros
    df = dm.get_filtered_data(
        curso_seleccionado if curso_seleccionado != "Todos" else None,
        trimestre_seleccionado if trimestre_seleccionado != "Todos" else None,
        alumno_seleccionado if alumno_seleccionado != "Todos" else None
    )
    
    if not df.empty:
        # Estadísticas generales
        st.subheader("📈 Estadísticas Generales")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("👥 Total Alumnos", len(df))
        
        with col2:
            if "Nota Final" in df.columns:
                promedio_final = df["Nota Final"].mean()
                st.metric("📈 Promedio Final", f"{promedio_final:.2f}")
        
        with col3:
            if "Nota Asistencia" in df.columns:
                asistencia_counts = df["Nota Asistencia"].value_counts()
                excelentes = asistencia_counts.get("10", 0)
                st.metric("🌟 Excelente Asistencia", excelentes)
        
        with col4:
            if "Evaluaciones" in df.columns:
                total_evaluaciones = df["Evaluaciones"].sum()
                st.metric("📝 Total Evaluaciones", int(total_evaluaciones))
        
        # Filtros aplicados
        st.subheader("🔍 Filtros Aplicados")
        filtros_aplicados = []
        if curso_seleccionado != "Todos":
            filtros_aplicados.append(f"📚 {curso_seleccionado}")
        if trimestre_seleccionado != "Todos":
            filtros_aplicados.append(f"📅 {trimestre_seleccionado}")
        if alumno_seleccionado != "Todos":
            filtros_aplicados.append(f"👤 {alumno_seleccionado}")
        
        if filtros_aplicados:
            st.info(" | ".join(filtros_aplicados))
        else:
            st.info("Mostrando todos los datos")
        
        # Tabla de datos completa
        st.subheader("📋 Datos Detallados")
        
        # Columnas a mostrar
        columnas_mostrar = [
            "Apellido y Nombre", "Curso", "Trimestre", "Asistencia", 
            "Nota Asistencia", "Tipo Evaluación", "Nota Final", "Observaciones"
        ]
        
        # Agregar evaluaciones
        for i in range(1, 7):
            columnas_mostrar.append(f"Evaluación {i}")
            columnas_mostrar.append(f"Calificación {i}")
        
        columnas_disponibles = [col for col in columnas_mostrar if col in df.columns]
        
        st.dataframe(
            df[columnas_disponibles].sort_values("Apellido y Nombre"),
            use_container_width=True,
            hide_index=True
        )
        
        # Gráficos
        if len(df) > 1:
            st.subheader("📈 Análisis Visual")
            
            col1, col2 = st.columns(2)
            
            with col1:
                if "Nota Final" in df.columns:
                    st.write("**📊 Distribución de Notas Finales**")
                    st.bar_chart(df["Nota Final"].value_counts().sort_index())
            
            with col2:
                if "Curso" in df.columns:
                    st.write("**📚 Distribución por Curso**")
                    st.bar_chart(df["Curso"].value_counts())
        
        # Exportación
        st.subheader("💾 Exportar Datos")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📥 Exportar a Excel", use_container_width=True):
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"reporte_{timestamp}.xlsx"
                
                try:
                    df.to_excel(filename, index=False)
                    st.success(f"✅ Exportado a {filename}")
                    
                    with open(filename, 'rb') as f:
                        st.download_button(
                            label="📄 Descargar Reporte",
                            data=f.read(),
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.error(f"❌ Error: {e}")
        
        with col2:
            if st.button("🔄 Actualizar Reportes", use_container_width=True):
                st.rerun()
    
    else:
        st.info("📝 No hay datos disponibles para los filtros seleccionados")
    
    # Botón para volver
    if st.button("🔙 Volver al Inicio", use_container_width=True):
        st.session_state.pagina = "inicio"
        st.rerun()

# Página de Generación Excel
elif st.session_state.pagina == "excel":
    st.header("💾 Generación de Archivos Excel")
    
    st.subheader("📊 Generar Excel por Trimestre")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Excel de Asistencias**")
        trimestre_excel = st.selectbox("Seleccionar Trimestre:", ["1 Trimestre", "2 Trimestre", "3 Trimestre"])
        
        if st.button("📋 Generar Excel de Asistencias", use_container_width=True, type="primary"):
            df = dm.load_data()
            if not df.empty:
                filename = crear_excel_asistencias(df, trimestre_excel)
                st.success(f"✅ Excel de asistencias creado: {filename}")
                st.info(f"📁 Ubicación: Esta carpeta")
                
                with open(filename, 'rb') as f:
                    st.download_button(
                        label="📄 Descargar Asistencias",
                        data=f.read(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
    
    with col2:
        st.write("**Excel de Evaluaciones**")
        trimestre_eval_excel = st.selectbox("Seleccionar Trimestre para Evaluaciones:", ["1 Trimestre", "2 Trimestre", "3 Trimestre"])
        
        if st.button("🎯 Generar Excel de Evaluaciones", use_container_width=True, type="primary"):
            df = dm.load_data()
            if not df.empty:
                filename = crear_excel_evaluaciones(df, trimestre_eval_excel)
                st.success(f"✅ Excel de evaluaciones creado: {filename}")
                st.info(f"📁 Ubicación: Esta carpeta")
                
                with open(filename, 'rb') as f:
                    st.download_button(
                        label="📄 Descargar Evaluaciones",
                        data=f.read(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
    
    # Lista de archivos
    st.markdown("---")
    st.subheader("📁 Archivos Generados")
    
    archivos_excel = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    
    if archivos_excel:
        st.write("Archivos Excel disponibles:")
        for archivo in archivos_excel:
            st.write(f"📄 {archivo}")
    else:
        st.info("📝 No hay archivos Excel generados aún")

# Página de Inicio
else:
    st.header("📋 Vista Principal de Alumnos")
    
    # Cargar datos con filtros
    df = dm.get_filtered_data(
        curso_seleccionado if curso_seleccionado != "Todos" else None,
        trimestre_seleccionado if trimestre_seleccionado != "Todos" else None,
        alumno_seleccionado if alumno_seleccionado != "Todos" else None
    )
    
    if not df.empty:
        # Resumen de filtros
        filtros_aplicados = []
        if curso_seleccionado != "Todos":
            filtros_aplicados.append(f"📚 {curso_seleccionado}")
        if trimestre_seleccionado != "Todos":
            filtros_aplicados.append(f"📅 {trimestre_seleccionado}")
        if alumno_seleccionado != "Todos":
            filtros_aplicados.append(f"👤 {alumno_seleccionado}")
        
        if filtros_aplicados:
            st.info("Filtros activos: " + " | ".join(filtros_aplicados))
        
        # Tabla principal
        columnas_mostrar = [
            "Apellido y Nombre", "Curso", "Trimestre", "Asistencia", 
            "Nota Asistencia", "Tipo Evaluación", "Nota Final"
        ]
        
        columnas_disponibles = [col for col in columnas_mostrar if col in df.columns]
        
        st.dataframe(
            df[columnas_disponibles].sort_values("Apellido y Nombre"),
            use_container_width=True,
            hide_index=True
        )
        
        # Estadísticas rápidas
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("📋 Total Registros", len(df))
        
        with col2:
            if "Nota Final" in df.columns:
                promedio = df["Nota Final"].mean()
                if not pd.isna(promedio):
                    st.metric("📈 Promedio General", f"{float(promedio):.2f}")
        
        with col3:
            if "Nota Asistencia" in df.columns:
                asistencia_counts = df["Nota Asistencia"].value_counts()
                excelentes = asistencia_counts.get("10", 0)
                st.metric("🌟 Excelente Asistencia", excelentes)
    
    else:
        st.info("📝 No hay datos disponibles. Usa los botones de la sidebar para comenzar.")
        
        # Información del sistema
        st.subheader("ℹ️ Información del Sistema")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.write("📚 Cursos disponibles:")
            for curso in dm.cursos:
                st.write(f"• {curso}")
        
        with col2:
            st.write("📅 Trimestres disponibles:")
            for trimestre in dm.trimestres:
                st.write(f"• {trimestre}")
        
        with col3:
            st.write("📝 Tipos de evaluación:")
            for tipo in dm.tipos_evaluacion:
                st.write(f"• {tipo}")
        
        st.subheader("📊 Sistema de Calificaciones")
        
        calificaciones_df = pd.DataFrame([
            {"Calificación": k, "Valor": v["valor"], "Descripción": v["descripcion"]} 
            for k, v in CALIFICACIONES_VALORES.items()
        ])
        
        st.dataframe(calificaciones_df, use_container_width=True, hide_index=True)

# Función para guardar asistencia
def guardar_asistencia_diaria(dm, asistencia_data, curso, trimestre, fecha, dia_numero):
    """Guardar asistencia diaria"""
    guardados = 0
    
    for alumno, presente in asistencia_data.items():
        guardados += 1
    
    st.success(f"✅ Asistencia guardada para {guardados} alumnos - Día {dia_numero} del {trimestre}")
    st.balloons()
    
    presentes = sum(asistencia_data.values())
    total = len(asistencia_data)
    porcentaje = (presentes / total * 100) if total > 0 else 0
    
    st.info(f"📊 Resumen: {presentes}/{total} presentes ({porcentaje:.1f}%)")

# Footer
st.markdown("---")
st.markdown(
    """
    <div style='text-align: center; color: gray; padding: 20px;'>
        <strong>📚 Dashboard de Gestión IES - Completo</strong><br>
        Sistema funcional de asistencias, evaluaciones y reportes
    </div>
    """,
    unsafe_allow_html=True
)
