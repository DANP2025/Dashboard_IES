import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class DataManagement:
    def __init__(self, excel_path="datos_alumnos.xlsx"):
        self.excel_path = excel_path
        self.cursos = ["EF 1A", "EF 2A", "EF 1B", "EF 2B", "TD 2A", "TD 2B"]
        self.trimestres = ["1 Trimestre", "2 Trimestre", "3 Trimestre"]
        self.tipos_evaluacion = ["Diagnóstico", "Físico", "Técnico", "Desempeño global"]
        self.calificaciones = {
            "M": {"rango": "Menos de 5", "valor": 4},
            "R-": {"rango": "5.5 - 6", "valor": 5.5},
            "R+": {"rango": "6", "valor": 6},
            "B": {"rango": "7", "valor": 7},
            "MB": {"rango": "8", "valor": 8},
            "Ex": {"rango": "9 - 10", "valor": 9.5}
        }
        self.initialize_excel()
    
    def initialize_excel(self):
        """Inicializa el archivo Excel si no existe"""
        if not os.path.exists(self.excel_path):
            self.create_excel_structure()
    
    def create_excel_structure(self):
        """Crea la estructura inicial del Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Generales"
        
        # Encabezados principales
        headers = [
            "Curso", "Trimestre", "Apellido y Nombre", "Asistencia",
            "Nota Asistencia", "Evaluaciones", "Tipo Evaluación",
            "Evaluación 1", "Calificación 1", "Evaluación 2", "Calificación 2",
            "Evaluación 3", "Calificación 3", "Evaluación 4", "Calificación 4",
            "Evaluación 5", "Calificación 5", "Evaluación 6", "Calificación 6",
            "Nota Final", "Observaciones", "Fecha Registro"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar ancho de columnas
        column_widths = [10, 12, 25, 15, 15, 15, 15, 15, 12, 15, 12, 15, 12, 15, 12, 15, 12, 15, 12, 12, 30, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        wb.save(self.excel_path)
    
    def load_data(self):
        """Carga los datos del Excel"""
        try:
            df = pd.read_excel(self.excel_path)
            return df
        except:
            return pd.DataFrame()
    
    def save_student_data(self, student_data):
        """Guarda los datos de un nuevo alumno"""
        try:
            df = self.load_data()
            new_row = pd.DataFrame([student_data])
            df = pd.concat([df, new_row], ignore_index=True)
            df.to_excel(self.excel_path, index=False)
            return True
        except Exception as e:
            print(f"Error al guardar datos: {e}")
            return False
    
    def calculate_attendance_grade(self, attendance_percentage):
        """Calcula la nota de asistencia según el porcentaje"""
        if attendance_percentage >= 80:
            return "Ex", 10
        elif attendance_percentage >= 51:
            return "MB", 8
        else:
            return "M", 5
    
    def calculate_final_grade(self, grades):
        """Calcula la nota final basada en las calificaciones"""
        if not grades:
            return 0
        
        numeric_grades = []
        for grade in grades:
            if grade in self.calificaciones:
                numeric_grades.append(self.calificaciones[grade]["valor"])
        
        if numeric_grades:
            return round(sum(numeric_grades) / len(numeric_grades), 2)
        return 0
    
    def get_filtered_data(self, curso=None, trimestre=None, alumno=None):
        """Filtra los datos según los criterios seleccionados"""
        df = self.load_data()
        
        if curso and curso != "Todos":
            df = df[df["Curso"] == curso]
        
        if trimestre and trimestre != "Todos":
            df = df[df["Trimestre"] == trimestre]
        
        if alumno and alumno != "Todos":
            df = df[df["Apellido y Nombre"] == alumno]
        
        return df
    
    def get_students_by_course(self, curso):
        """Obtiene la lista de alumnos de un curso específico"""
        df = self.load_data()
        if curso and curso != "Todos":
            df = df[df["Curso"] == curso]
        return sorted(df["Apellido y Nombre"].unique()) if not df.empty else []
